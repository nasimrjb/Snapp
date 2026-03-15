"""
header_frequency.py  —  Utility / Diagnostic Script
=====================================================

PURPOSE:
    This is a diagnostic tool that checks how often each survey column header
    appears across all raw Excel files.  It compares the headers found in the
    raw data against the list of known questions in column_rename.xlsx, and
    produces a formatted Excel report showing:
      • Which known questions were found (and how many times)
      • Which raw headers are NOT yet filed in the rename mapping
      • Which known questions have 0 occurrences (stale/removed questions)

WHY IT EXISTS:
    New survey waves sometimes add, remove, or rename questions.  Running this
    script after receiving new raw data tells you immediately what's new and
    what might be missing from the column_rename.xlsx mapping.

HOW TO RUN:
    python header_frequency.py

OUTPUT:
    DataSources/header_frequency_report.xlsx  (two sheets: "Header Frequency"
    and "Summary")

PIPELINE POSITION:
    This script is NOT part of the main pipeline.  It's a standalone diagnostic
    you run when you receive new data to audit whether the mapping is up to date.
"""

import os
import re
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIGURATION ────────────────────────────────────────────────────────────
# Where raw survey files live, where the rename Excel is, and where to save output
RAW_DIR = r"D:\Work\Driver Survey\raw"
RENAME_FILE = r"D:\Work\Driver Survey\DataSources\column_rename.xlsx"
OUTPUT_FILE = r"D:\Work\Driver Survey\DataSources\header_frequency_report.xlsx"

# ── Persian / Arabic text normalisation ──────────────────────────────────────
# Survey headers can contain Arabic look-alike characters (e.g. Arabic "ك" vs
# Persian "ک").  This translation table converts Arabic forms to their Persian
# equivalents, and strips invisible Unicode characters (zero-width non-joiner,
# zero-width space, right-to-left mark, byte-order mark, etc.).
# This mirrors the normalisation logic used in the main data_cleaning pipeline,
# so that matching is consistent.
_AR_TO_FA = str.maketrans({
    "\u0643": "\u06a9",  # Arabic kaf → Persian kaf
    "\u064a": "\u06cc",  # Arabic yeh → Persian yeh
    "\u0629": "\u0647",  # Arabic teh marbuta → Persian heh
    "\u0624": "\u0648",  # Arabic waw w/ hamza → plain waw
    "\u0625": "\u0627",  # Arabic alef w/ hamza below → plain alef
    "\u0623": "\u0627",  # Arabic alef w/ hamza above → plain alef
    "\u0671": "\u0627",  # Arabic alef wasla → plain alef
    "\u200c": " ",       # ZWNJ (zero-width non-joiner) → space
    "\u200b": "",        # zero-width space → remove
    "\u200f": "",        # right-to-left mark → remove
    "\u200e": "",        # left-to-right mark → remove
    "\ufeff": "",        # byte-order mark → remove
})

# Regex pattern to strip Arabic/Persian diacritical marks (short vowels, etc.)
_DIACRITICS = re.compile(
    r"[\u064b-\u065f\u0610-\u061a\u06d6-\u06dc\u06df-\u06e4\u06e7\u06e8\u06ea-\u06ed]"
)


def fa_norm(text: str) -> str:
    """
    Normalize a Persian/Arabic string for comparison.

    Steps:
      1. Convert Arabic look-alike chars to their Persian equivalents
      2. Strip diacritical marks (short vowels, tanwin, etc.)
      3. Collapse multiple whitespace characters into a single space
      4. Strip leading/trailing whitespace

    This ensures that two headers that look visually identical but differ in
    invisible Unicode characters will match each other.
    """
    text = str(text).translate(_AR_TO_FA)
    text = _DIACRITICS.sub("", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1: Read all headers from every raw .xlsx file
# ══════════════════════════════════════════════════════════════════════════════
raw_path = Path(RAW_DIR)
xlsx_files = sorted(raw_path.glob("*.xlsx"))
print(f"Found {len(xlsx_files)} .xlsx files in {RAW_DIR}")

# Counter tracks how many times each normalised header appears across all files.
# We also remember the first "raw" (un-normalised) form we saw, and which
# files contained each header, for the report.
header_counter = Counter()          # normalised header -> count
header_raw_examples = {}            # normalised header -> first raw text seen
files_with_header = {}              # normalised header -> set of filenames

for fpath in xlsx_files:
    fname = fpath.name
    # Skip Excel temp files (these are created while a file is open in Excel)
    if fname.startswith("~$"):
        continue
    try:
        # Open the workbook and iterate over all sheets
        xls = pd.ExcelFile(fpath, engine="openpyxl")
        for sheet in xls.sheet_names:
            # nrows=0 means "read zero data rows" — we only want the header row
            df = pd.read_excel(xls, sheet_name=sheet, nrows=0)
            for col in df.columns:
                raw = str(col).strip()
                norm = fa_norm(raw)
                if not norm:
                    continue
                header_counter[norm] += 1
                # Save the first raw text we saw for this normalised header
                if norm not in header_raw_examples:
                    header_raw_examples[norm] = raw
                # Track which files contained this header
                files_with_header.setdefault(norm, set()).add(fname)
    except Exception as e:
        print(f"  ⚠ Skipped {fname}: {e}")

print(f"Found {len(header_counter)} unique headers (after normalisation)")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 2: Load known questions from column_rename.xlsx
# ══════════════════════════════════════════════════════════════════════════════
# This is the same file that generate_mapping.py reads.  We pull out the
# question_raw (original Persian header) and question_short (clean English name)
# columns to compare against what we found in the raw files.
qs = pd.read_excel(RENAME_FILE, sheet_name="questions")
known_raw_col = "question_raw"
known_short_col = "question_short"

known_questions_raw = qs[known_raw_col].dropna().astype(
    str).str.strip().tolist()
known_questions_short = qs[known_short_col].tolist()
known_questions_norm = [fa_norm(q) for q in known_questions_raw]

# Build ordered lookup maps so we can go from normalised → raw text / short name
known_map = {}        # normalised → raw text
known_short_map = {}  # normalised → question_short
for raw, short, norm in zip(known_questions_raw, known_questions_short, known_questions_norm):
    if norm not in known_map:
        known_map[norm] = raw
        known_short_map[norm] = str(short).strip() if pd.notna(short) else ""

print(f"Loaded {len(known_map)} known questions from column_rename.xlsx")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 3: Build the output rows
# ══════════════════════════════════════════════════════════════════════════════
# Each row will be: (question_text, short_name, frequency, status)
# Status is either "Available" (in the rename mapping) or "Not filed" (new/unknown).
rows = []
seen_norms = set()

# First pass: all questions from column_rename in their original Excel order
# This ensures the report follows the same order as the mapping file.
for norm in known_questions_norm:
    if norm in seen_norms:
        continue
    seen_norms.add(norm)
    freq = header_counter.get(norm, 0)
    raw_text = known_map[norm]
    short_name = known_short_map.get(norm, "")
    rows.append((raw_text, short_name, freq, "Available"))

# Second pass: headers found in raw files but NOT in column_rename
# These are new questions that haven't been added to the mapping yet.
for norm, count in header_counter.most_common():
    if norm not in seen_norms:
        seen_norms.add(norm)
        raw_text = header_raw_examples[norm]
        rows.append((raw_text, "", count, "Not filed"))

# ══════════════════════════════════════════════════════════════════════════════
# STEP 4: Create a formatted Excel report using openpyxl
# ══════════════════════════════════════════════════════════════════════════════
# openpyxl lets us create .xlsx files with custom formatting (colours, fonts,
# borders, etc.) — much nicer than a plain CSV.
wb = Workbook()
ws = wb.active
ws.title = "Header Frequency"

# --- Define reusable styles ------------------------------------------------
# These are openpyxl style objects we apply to cells to make the report look nice.
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F5496")  # dark blue background
header_align = Alignment(
    horizontal="center", vertical="center", wrap_text=True)

avail_fill = PatternFill("solid", fgColor="E2EFDA")     # light green for "Available"
not_filed_fill = PatternFill("solid", fgColor="FCE4EC")  # light red for "Not filed"
zero_font = Font(name="Arial", color="C00000", size=10)  # red text for 0 frequency
normal_font = Font(name="Arial", size=10)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# --- Write header row (row 1) ----------------------------------------------
headers = ["Question", "Question Short", "Frequency", "Status"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# --- Write data rows (row 2 onwards) ---------------------------------------
for r, (question, short_name, freq, status) in enumerate(rows, 2):
    c1 = ws.cell(row=r, column=1, value=question)
    c2 = ws.cell(row=r, column=2, value=short_name)
    c3 = ws.cell(row=r, column=3, value=freq)
    c4 = ws.cell(row=r, column=4, value=status)

    # Apply fonts and alignment
    c1.font = normal_font
    c1.alignment = Alignment(vertical="center", wrap_text=True)
    c2.font = normal_font
    c2.alignment = Alignment(vertical="center")
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c4.alignment = Alignment(horizontal="center", vertical="center")

    # Highlight frequency = 0 in red (means this question is in the mapping
    # but was never found in any raw file — possibly removed from the survey)
    if freq == 0:
        c3.font = zero_font
    else:
        c3.font = normal_font

    # Color-code the status column
    if status == "Available":
        c4.fill = avail_fill
        c4.font = Font(name="Arial", size=10, color="375623")  # dark green
    else:
        c4.fill = not_filed_fill
        c4.font = Font(name="Arial", size=10, color="C00000")  # red

    # Apply thin borders to all cells in the row
    for cell in (c1, c2, c3, c4):
        cell.border = thin_border

# --- Set column widths for readability --------------------------------------
ws.column_dimensions["A"].width = 65   # Question text (Persian, can be long)
ws.column_dimensions["B"].width = 30   # Question short name
ws.column_dimensions["C"].width = 14   # Frequency
ws.column_dimensions["D"].width = 14   # Status

# Freeze the header row so it stays visible when scrolling
ws.freeze_panes = "A2"

# Add auto-filter (the dropdown arrows on each column header in Excel)
ws.auto_filter.ref = f"A1:D{len(rows) + 1}"

# ── Summary sheet ────────────────────────────────────────────────────────────
# A second sheet with high-level statistics
ws2 = wb.create_sheet("Summary")
summary_data = [
    ("Total .xlsx files scanned", len(xlsx_files)),
    ("Unique headers found", len(header_counter)),
    ("Known questions (column_rename)", len(known_map)),
    ("Available (matched)", sum(1 for _, _, _, s in rows if s == "Available")),
    ("Not filed (new headers)", sum(1 for _, _, _, s in rows if s == "Not filed")),
    ("Available with 0 frequency", sum(1 for _, _, f,
     s in rows if s == "Available" and f == 0)),
]
for r, (label, val) in enumerate(summary_data, 1):
    ws2.cell(row=r, column=1, value=label).font = Font(
        name="Arial", bold=True, size=10)
    ws2.cell(row=r, column=2, value=val).font = Font(name="Arial", size=10)
    ws2.cell(row=r, column=2).alignment = Alignment(horizontal="center")
ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 14

# --- Save the workbook ------------------------------------------------------
wb.save(OUTPUT_FILE)
print(f"\n✅ Report saved to: {OUTPUT_FILE}")
print(f"   {sum(1 for _, _, _, s in rows if s == 'Available')} Available | "
      f"{sum(1 for _, _, _, s in rows if s == 'Not filed')} Not filed | "
      f"{sum(1 for _, _, f, s in rows if s == 'Available' and f == 0)} Available with 0 frequency")
