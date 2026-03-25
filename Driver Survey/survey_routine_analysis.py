"""
Driver Survey Routine Analysis
Produces conceptually similar outputs to the weekly Excel routine report.
Reads from the 6 processed CSV files and generates analysis tables by city/week.
"""

import pandas as pd
import numpy as np
import os
import warnings
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

warnings.filterwarnings("ignore")

# ─── Config ──────────────────────────────────────────────────────────────────
BASE_DIR = r"D:\Work\Driver Survey"
PROCESSED_DIR = os.path.join(BASE_DIR, "processed")
OUTPUT_DIR = os.path.join(BASE_DIR, "processed")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Ordered list of top cities (display order preserved in all outputs)
TOP_CITIES = [
    "Tehran(city)", "Karaj", "Isfahan", "Shiraz", "Mashhad", "Qom",
    "Tabriz", "Ahwaz", "Sari", "Rasht", "Urumieh", "Yazd",
    "Kerman", "Gorgan", "Ghazvin", "Arak", "Kermanshah", "Hamedan",
    "Ardebil", "Bojnurd", "Khorramabad", "Zanjan", "Kish",
]

MERGE_COLS = ["recordID", "city", "active_time", "cooperation_type",
              "age_group", "edu", "marr_stat", "gender", "original_job"]

# ─── Column Ordering Maps ────────────────────────────────────────────────────
INCENTIVE_AMOUNT_ORDER = [
    "< 100k", "100_200k", "100_250k", "200_400k", "250_500k",
    "400_600k", "500_750k", "600_800k", "750k_1m", "800k_1m",
    "1m_1.25m", "1.25m_1.5m", ">1.5m",
]
INCENTIVE_DURATION_ORDER = [
    "Few Hours", "1 Day", "1_6 Days", "7 Days", ">7 Days",
]
ACTIVITY_TYPE_ORDER = [
    "few hours/month", "<20hour/mo", "5_20hour/week",
    "20_40h/week", ">40h/week", "8_12hour/day", ">12h/day",
]
INACTIVITY_ORDER = [
    "Same Day", "1_3 Day Before", "3_7 Days Before",
    "8_14 Days Before", "15_30 Days_Before",
    "1_2 Month Before", "2_3 Month Before",
    "3_6Month Before", ">6 Month Before",
]

COLUMN_ORDERS = {
    "#1_Snapp_Incentive_Amt": INCENTIVE_AMOUNT_ORDER,
    "#2_Tapsi_Incentive_Amt": INCENTIVE_AMOUNT_ORDER,
    "#4_Incentive_Duration": (
        [f"Snapp_{v}" for v in INCENTIVE_DURATION_ORDER] + ["Snapp_n"]
        + [f"Tapsi_{v}" for v in INCENTIVE_DURATION_ORDER] + ["Tapsi_n"]
    ),
    "#15_Persona_Activity Type": ACTIVITY_TYPE_ORDER,
    "#17_Inactivity": INACTIVITY_ORDER,
}

# ─── Minimum Sample-Size Cutoffs ─────────────────────────────────────────
# Per-sheet prefix → minimum count required for a row to show data.
# Rows (cities) below the threshold will have their data blanked.
#
# Value format:
#   -1           → disabled (no filtering)
#   5            → filter rows where the primary count col (n, respondent_count, …) < 5
#   {"n": 5, "n_joint": 3}  → filter independently per count column;
#       columns whose names start with "Jnt_" are blanked by n_joint,
#       all other data columns are blanked by n.
SHEET_MIN_N = {
    "#1_Snapp_Incentive": 5,
    "#2_Tapsi_Incentive": 5,
    "#3_Sat": {"n": 17, "n_joint": -1},
    "#4_Incentive_Duration": {"Snapp_n": -1, "Tapsi_n": -1},
    "#5_6_IncType": {"ExSnapp_n": -1, "JntSnapp_n": -1, "Tapsi_n": -1},
    "#8_Dissat": {"Snapp_n": 10, "Tapsi_n": 10},
    "#9_Dissat_Sum": -1,
    "#12_Cities_Overview": {"E_n": 17, "F_n": 17, "G_n": 17},
    "#13_RideShare": -1,
    "#14_Nav": -1,
    "#15_Persona": -1,
    "#16_Ref": -1,
    "#17_Inactivity": -1,
    "#18_CommFree": -1,
    "#19_LuckyWheel": -1,
    "#20_Refusal": -1,
    "#CS_Sat": -1,
    "#CS_Cat": -1,
    "#CS_Reason": -1,
    "#Reco_NPS": -1,
    "#Refer": -1,
    "#NavReco": -1,
    "#Reg": -1,
    "#Income": -1,
    "#Decline": -1,
    "#Carfix": -1,
    "#Garage": -1,
    "#Demand": -1,
    "#Speed": -1,
    "#DistOrigin": -1,
    "#GPS": -1,
    "#Unpaid": -1,
}

# Columns that should NOT be treated as percentage
NON_PCT_COLS = {"n", "n_joint", "Snapp_n", "Tapsi_n", "ExSnapp_n", "JntSnapp_n",
                "n_dissatisfied", "n_contacted",
                "respondent_count", "joint_count",
                "avg_snapp_LOC", "avg_tapsi_LOC",
                "avg_snapp_ride", "avg_tapsi_ride",
                "avg_magical_window_income", "count",
                "tapsi_carpooling_count",
                "E_n", "F_n", "G_n",
                "AvgLOC_All Snapp", "AvgLOC_Joint Snapp",
                "AvgLOC_Joint Cmpt", "AvgLOC_Joint Cmpt SU",
                "total Res", "Joint Res", "Ex drivers",
                "Total Ride", "Total Ride Snapp",
                "Ex drivers Ride in Snapp",
                "Snapp Ride", "Tapsi Tide",
                "RidePerBoarded_Snapp", "RidePerBoarded_Tapsi",
                "AvgAllRides",
                "Who Got Message", "Free Comm Drivers",
                "GotMsg_Money", "GotMsg_Free-Commission",
                "GotMsg_Money & Free-commission",
                "Total Rides",
                "RidesAmong_Total Rides", "RidesAmong_Free Comm Rides"}

# Count columns used for min-n cutoff detection (order = priority)
_COUNT_COLS_PRIORITY = ["n", "n_joint", "Snapp_n", "Tapsi_n", "ExSnapp_n", "JntSnapp_n",
                        "E_n", "F_n", "G_n", "respondent_count",
                        "n_dissatisfied", "n_contacted", "count"]

# Sheet prefixes where values are satisfaction scores (1-5), not percentages
SATISFACTION_SHEETS = {"#3_Sat_", "#CS_Sat_", "#Carfix_Sat_", "#Garage_Sat_",
                       "#NavReco_"}

# Sheets where values are absolute numbers, not percentages
ABSOLUTE_SHEETS = {"#Demand_"}


# ─── Helpers ─────────────────────────────────────────────────────────────────
def load_data():
    print("Loading data...")
    data = {}
    files = {
        "short_main": "short_survey_main.csv",
        "short_rare": "short_survey_rare.csv",
        "wide_main": "wide_survey_main.csv",
        "wide_rare": "wide_survey_rare.csv",
        "long_main": "long_survey_main.csv",
        "long_rare": "long_survey_rare.csv",
    }
    for key, fname in files.items():
        path = os.path.join(PROCESSED_DIR, fname)
        print(f"  Loading {fname}...")
        if not os.path.exists(path):
            print(f"    WARNING: {fname} not found, using empty DataFrame")
            data[key] = pd.DataFrame()
            continue
        try:
            data[key] = pd.read_csv(path, low_memory=False)
        except pd.errors.EmptyDataError:
            print(f"    WARNING: {fname} is empty, using empty DataFrame")
            data[key] = pd.DataFrame()
    print(f"  Done. short_main shape: {data['short_main'].shape}")
    available = [c for c in MERGE_COLS if c in data["short_main"].columns]
    data["_lookup"] = data["short_main"][available].drop_duplicates(
        subset="recordID")
    return data


def get_latest_week(df, min_respondents=100):
    counts = df.groupby("weeknumber").size()
    valid = counts[counts >= min_respondents]
    return valid.index.max() if len(valid) > 0 else df["weeknumber"].max()


def filter_week(df, week):
    return df[df["weeknumber"] == week].copy()


def filter_top_cities(df, city_col="city"):
    if city_col not in df.columns:
        return df
    return df[df[city_col].isin(TOP_CITIES)]


def add_city(df, lookup):
    if "city" in df.columns:
        return df
    merge_cols = [
        c for c in lookup.columns if c not in df.columns or c == "recordID"]
    return df.merge(lookup[merge_cols], on="recordID", how="left")


def sort_cities(df):
    """Reindex rows to match the TOP_CITIES display order."""
    if df.index.name == "City" or (hasattr(df.index, 'name') and df.index.name is None):
        # Keep cities that exist in data, in TOP_CITIES order, then any extras
        ordered = [c for c in TOP_CITIES if c in df.index]
        extras = [c for c in df.index if c not in TOP_CITIES and c != "Total"]
        has_total = "Total" in df.index
        new_order = ordered + extras + (["Total"] if has_total else [])
        return df.reindex(new_order)
    return df


def add_total_row(ct, n_total=None):
    total = ct.select_dtypes(include="number").mean().to_frame().T
    total.index = ["Total"]
    if n_total is not None and "n" in ct.columns:
        total["n"] = n_total
    result = pd.concat([ct, total]).round(4)
    return sort_cities(result)


def reorder_columns(df, sheet_name):
    order = COLUMN_ORDERS.get(sheet_name)
    if order is None:
        return df
    ordered = [c for c in order if c in df.columns]
    others = [c for c in df.columns if c not in order]
    return df[ordered + others]


def get_min_n(sheet_name):
    """Look up the minimum sample-size cutoff for a sheet by prefix match.

    Returns int or dict (e.g. {"n": 5, "n_joint": 3}).
    """
    for prefix, threshold in SHEET_MIN_N.items():
        if sheet_name.startswith(prefix):
            return threshold
    return -1


def apply_min_n_cutoff(df, sheet_name):
    """Blank data columns for rows where the count is below the sheet threshold.

    The count column(s) themselves are preserved so the analyst can see
    why a row is blank.  The 'Total' row is never filtered.

    When threshold is a dict (e.g. {"n": 5, "n_joint": 3}):
      - Columns containing "Jnt_" in their name are blanked by n_joint threshold
      - All other data columns are blanked by n threshold
    When threshold is a plain int:
      - All data columns are blanked by the first available count column
    """
    raw = get_min_n(sheet_name)
    if df.empty:
        return df

    # Dict mode: independent cutoffs per count column
    if isinstance(raw, dict):
        active = {k: v for k, v in raw.items() if v > 0 and k in df.columns}
        if not active:
            return df
        df = df.copy()
        not_total = df.index != "Total"
        count_set = set(NON_PCT_COLS)
        data_cols = [c for c in df.columns if c not in count_set]

        # Derive prefix/substring each count column controls:
        #   "n_joint" → "Jnt_",  "Snapp_n" → "Snapp_",  "Tapsi_n" → "Tapsi_"
        #   "ExSnapp_n" → "Exclusives Snapp",  "JntSnapp_n" → "Joints Snapp"
        #   "n" → None (fallback: everything not matched by other prefixes)
        prefixes = {}
        for k in active:
            if k == "n_joint":
                prefixes[k] = "Jnt_"
            elif k == "ExSnapp_n":
                prefixes[k] = "Exclusives Snapp"
            elif k == "JntSnapp_n":
                prefixes[k] = "Joints Snapp"
            elif k == "E_n":
                # #12 Cities: E controls % Joint, % Dual SU, AvgLOC_All, GotMsg_All
                prefixes[k] = "_E_"  # matched via _CITIES_CUTOFF_MAP
            elif k == "F_n":
                prefixes[k] = "_F_"
            elif k == "G_n":
                prefixes[k] = "_G_"
            elif k.endswith("_n") and k != "n":
                prefixes[k] = k[:-1]  # "Snapp_n" → "Snapp_"
            else:
                prefixes[k] = None  # fallback

        # #12 Cities Overview: explicit column-group mapping for E/F/G cutoffs
        # (includes NON_PCT_COLS like AvgLOC_ that must also be blanked)
        _cities_col_map = {
            "_E_": {"% Joint", "% Dual SU", "AvgLOC_All Snapp", "GotMsg_All Snapp"},
            "_F_": {"AvgLOC_Joint Snapp", "GotMsg_Joint Snapp", "GotMsg_Joint Cmpt"},
            "_G_": {"AvgLOC_Joint Cmpt", "AvgLOC_Joint Cmpt SU", "GotMsg_Joint Cmpt SU"},
        }

        all_cols = [c for c in df.columns if c not in active]
        for count_col, threshold in active.items():
            mask = (df[count_col] < threshold) & not_total
            prefix = prefixes[count_col]
            if prefix in _cities_col_map:
                # Explicit column set for #12 Cities Overview (uses all_cols, not data_cols)
                cols = [c for c in all_cols if c in _cities_col_map[prefix]]
            elif prefix is not None:
                # Match columns that start with prefix OR contain the substring
                cols = [c for c in data_cols
                        if c.startswith(prefix) or prefix in c]
            else:
                # Fallback: columns not claimed by any other prefix
                other_pfx = [p for k, p in prefixes.items()
                             if p is not None and k != count_col]
                cols = [c for c in data_cols
                        if not any(c.startswith(p) or p in c for p in other_pfx)]
            if cols:
                df.loc[mask, cols] = np.nan
        return df

    # Simple int mode
    threshold = raw
    if threshold <= 0:
        return df

    primary = None
    for c in _COUNT_COLS_PRIORITY:
        if c in df.columns:
            primary = c
            break
    if primary is None:
        return df

    count_set = set(NON_PCT_COLS)
    data_cols = [c for c in df.columns if c not in count_set]

    mask = (df[primary] < threshold) & (df.index != "Total")
    df = df.copy()
    df.loc[mask, data_cols] = np.nan
    return df


def is_pct_sheet(sheet_name):
    for prefix in SATISFACTION_SHEETS:
        if sheet_name.startswith(prefix):
            return False
    for name in ABSOLUTE_SHEETS:
        if sheet_name.startswith(name):
            return False
    return True


def convert_pct_to_decimal(df, sheet_name):
    if not is_pct_sheet(sheet_name):
        for c in df.columns:
            if (c.endswith("_%") or c.endswith("_pct")
                    or "Part%" in c or "GotMsg%" in c):
                df[c] = df[c] / 100
        return df
    for c in df.columns:
        if c in NON_PCT_COLS:
            continue
        if pd.api.types.is_numeric_dtype(df[c]):
            df[c] = df[c] / 100
    return df


def _get_sat_group_col_ranges(df):
    """For #3_Sat_ sheets, group columns by metric prefix for shared formatting."""
    groups = {}
    for prefix, _ in SAT_GROUP_HEADERS:
        col_indices = []
        for col_idx, col_name in enumerate(df.columns, start=2):
            if col_name.startswith(prefix):
                col_indices.append(col_idx)
        if col_indices:
            groups[prefix] = col_indices
    return groups


def apply_conditional_formatting(ws, df, sheet_name, data_start_row=2):
    if df.empty:
        return
    n_rows = len(df)
    is_sat_sheet = sheet_name.startswith("#3_Sat_")

    if is_sat_sheet:
        # Apply one color scale per metric group (all 3 columns share the range)
        groups = _get_sat_group_col_ranges(df)
        for prefix, col_indices in groups.items():
            is_pct = ("Part%" in prefix or "GotMsg%" in prefix)
            first_letter = get_column_letter(min(col_indices))
            last_letter = get_column_letter(max(col_indices))
            end_row = data_start_row + n_rows - 1
            cell_range = f"{first_letter}{data_start_row}:{last_letter}{end_row}"
            if is_pct:
                rule = ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    end_type="max", end_color="63BE7B",
                )
            else:
                rule = ColorScaleRule(
                    start_type="num", start_value=1, start_color="F8696B",
                    mid_type="num", mid_value=3, mid_color="FFEB84",
                    end_type="num", end_value=5, end_color="63BE7B",
                )
            ws.conditional_formatting.add(cell_range, rule)
        return

    # #12 Cities Overview: per-column conditional formatting matching reference
    if sheet_name.startswith("#12_Cities_Overview"):
        end_row = data_start_row + n_rows - 1
        # Column-specific color rules:
        #   % Joint → white→green,  % Dual SU → white→red
        #   AvgLOC_* → white→green (one shared range J:M)
        #   GotMsg_All Snapp → white→green (standalone)
        #   GotMsg_Joint* → white→green (shared range O:Q)
        _col_rules = {
            "% Joint":     "63BE7B",  # green
            "% Dual SU":   "F8696B",  # red
        }
        for col_idx, col_name in enumerate(df.columns, start=2):
            if col_name in _col_rules:
                cl = get_column_letter(col_idx)
                cr = f"{cl}{data_start_row}:{cl}{end_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FCFCFF",
                    end_type="max", end_color=_col_rules[col_name],
                )
                ws.conditional_formatting.add(cr, rule)
        # AvgLOC group (J:M in reference) — all white→green
        loc_cols = [i for i, c in enumerate(df.columns, start=2)
                    if c.startswith("AvgLOC_")]
        if loc_cols:
            fl = get_column_letter(min(loc_cols))
            ll = get_column_letter(max(loc_cols))
            cr = f"{fl}{data_start_row}:{ll}{end_row}"
            ws.conditional_formatting.add(cr, ColorScaleRule(
                start_type="min", start_color="FCFCFF",
                end_type="max", end_color="63BE7B"))
        # GotMsg_All Snapp standalone — white→green
        for col_idx, col_name in enumerate(df.columns, start=2):
            if col_name == "GotMsg_All Snapp":
                cl = get_column_letter(col_idx)
                cr = f"{cl}{data_start_row}:{cl}{end_row}"
                ws.conditional_formatting.add(cr, ColorScaleRule(
                    start_type="min", start_color="FCFCFF",
                    end_type="max", end_color="63BE7B"))
        # GotMsg_Joint* group (O:Q in reference) — white→green
        msg_joint_cols = [i for i, c in enumerate(df.columns, start=2)
                         if c.startswith("GotMsg_") and c != "GotMsg_All Snapp"]
        if msg_joint_cols:
            fl = get_column_letter(min(msg_joint_cols))
            ll = get_column_letter(max(msg_joint_cols))
            cr = f"{fl}{data_start_row}:{ll}{end_row}"
            ws.conditional_formatting.add(cr, ColorScaleRule(
                start_type="min", start_color="FCFCFF",
                end_type="max", end_color="63BE7B"))
        return

    # #13 RideShare: cellIs=0 pink highlight on count cols, per-pct-col color scales
    if sheet_name.startswith("#13_RideShare"):
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill, Font as OpFont
        end_row = data_start_row + n_rows - 1

        # Count columns: pink highlight when value=0
        _rs_counts = {"total Res", "Joint Res", "Ex drivers",
                      "Total Ride", "Total Ride Snapp",
                      "Ex drivers Ride in Snapp", "Snapp Ride", "Tapsi Tide"}
        cnt_idxs = [i for i, c in enumerate(df.columns, start=2) if c in _rs_counts]
        if cnt_idxs:
            fl = get_column_letter(min(cnt_idxs))
            ll = get_column_letter(max(cnt_idxs))
            cr = f"{fl}{data_start_row}:{ll}{end_row}"
            ws.conditional_formatting.add(cr, CellIsRule(
                operator="equal", formula=["0"],
                fill=PatternFill(bgColor="FFC7CE"),
                font=OpFont(color="9C0006")))

        # Percentage columns: white→green, except @Tapsi → white→red
        _rs_pct_colors = {
            "All Snapp": "63BE7B",
            "Ex Drivers in Snapp": "63BE7B",
            "Jnt @Snapp": "63BE7B",
            "Jnt @Tapsi": "F8696B",
        }
        for col_idx, col_name in enumerate(df.columns, start=2):
            if col_name in _rs_pct_colors:
                cl = get_column_letter(col_idx)
                cr = f"{cl}{data_start_row}:{cl}{end_row}"
                ws.conditional_formatting.add(cr, ColorScaleRule(
                    start_type="min", start_color="FCFCFF",
                    end_type="max", end_color=_rs_pct_colors[col_name]))
        return

    # Check if this sheet has group headers → apply one scale per group
    grp_headers = None
    for prefix, headers in SHEET_GROUP_HEADERS.items():
        if sheet_name.startswith(prefix):
            grp_headers = headers
            break

    if grp_headers is not None:
        # Dissatisfaction sheets use red scale (higher = worse)
        is_dissat = sheet_name.startswith("#8_Dissat")
        end_row = data_start_row + n_rows - 1
        for grp_prefix, _label in grp_headers:
            col_indices = [
                idx for idx, c in enumerate(df.columns, start=2)
                if c.startswith(grp_prefix) and c not in NON_PCT_COLS
                and pd.api.types.is_numeric_dtype(df[c])
            ]
            if not col_indices:
                continue
            first_letter = get_column_letter(min(col_indices))
            last_letter = get_column_letter(max(col_indices))
            cell_range = f"{first_letter}{data_start_row}:{last_letter}{end_row}"
            if is_dissat:
                rule = ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    end_type="max", end_color="F8696B",
                )
            else:
                rule = ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    end_type="max", end_color="63BE7B",
                )
            ws.conditional_formatting.add(cell_range, rule)
        return

    # Non-grouped sheets: per-column formatting
    is_dissat_summary = sheet_name.startswith("#9_Dissat")
    for col_idx, col_name in enumerate(df.columns, start=2):
        if col_name in NON_PCT_COLS:
            continue
        if not pd.api.types.is_numeric_dtype(df[col_name]):
            continue
        # Skip WoW columns from conditional formatting
        if "WoW" in str(col_name):
            continue
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}{data_start_row}:{col_letter}{data_start_row + n_rows - 1}"

        is_pct_col = ("Part%" in str(col_name)
                      or "GotMsg%" in str(col_name))

        if is_pct_col:
            rule = ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                end_type="max", end_color="63BE7B",
            )
        elif is_dissat_summary:
            rule = ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                end_type="max", end_color="F8696B",
            )
        elif any(sheet_name.startswith(p) for p in SATISFACTION_SHEETS):
            rule = ColorScaleRule(
                start_type="num", start_value=1, start_color="F8696B",
                mid_type="num", mid_value=3, mid_color="FFEB84",
                end_type="num", end_value=5, end_color="63BE7B",
            )
        else:
            rule = ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                end_type="max", end_color="63BE7B",
            )
        ws.conditional_formatting.add(cell_range, rule)


def format_pct_cells(ws, df, sheet_name, data_start_row=2):
    if df.empty:
        return

    # #12 Cities Overview: custom number formats per column type
    if sheet_name.startswith("#12_Cities_Overview"):
        for col_idx, col_name in enumerate(df.columns, start=2):
            if not pd.api.types.is_numeric_dtype(df[col_name]):
                continue
            if col_name.startswith("AvgLOC_"):
                fmt = '0.0'
            elif col_name in ("E_n", "F_n", "G_n"):
                continue  # keep General for counts
            else:
                fmt = '0%'  # % Joint, % Dual SU, GotMsg_*
            for row_idx in range(data_start_row, data_start_row + len(df)):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = fmt
        return

    # #18 CommFree: count cols as General, pct cols as 0%
    if sheet_name.startswith("#18_CommFree"):
        _cf_count_cols = {"n", "Who Got Message", "Free Comm Drivers",
                          "GotMsg_Money", "GotMsg_Free-Commission",
                          "GotMsg_Money & Free-commission",
                          "Total Rides", "RidesAmong_Total Rides",
                          "RidesAmong_Free Comm Rides"}
        for col_idx, col_name in enumerate(df.columns, start=2):
            if not pd.api.types.is_numeric_dtype(df[col_name]):
                continue
            if col_name in _cf_count_cols:
                continue  # keep General
            else:
                fmt = '0%'
            for row_idx in range(data_start_row, data_start_row + len(df)):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = fmt
        return

    # #15 Persona PartTime: counts/rides as General, pct cols as 0%, WoW as 0%
    if sheet_name.startswith("#15_Persona_PartTime"):
        _pt_count_cols = {"total Res", "Joint Res", "Ex drivers",
                          "RidePerBoarded_Snapp", "RidePerBoarded_Tapsi",
                          "AvgAllRides"}
        for col_idx, col_name in enumerate(df.columns, start=2):
            if not pd.api.types.is_numeric_dtype(df[col_name]):
                continue
            if col_name in _pt_count_cols:
                continue  # keep General
            else:
                fmt = '0%'
            for row_idx in range(data_start_row, data_start_row + len(df)):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = fmt
        return

    # #13 RideShare: counts as General, pct cols as 0%, WoW as 0.00%
    if sheet_name.startswith("#13_RideShare"):
        _rs_count_cols = {"total Res", "Joint Res", "Ex drivers",
                          "Total Ride", "Total Ride Snapp",
                          "Ex drivers Ride in Snapp", "Snapp Ride", "Tapsi Tide"}
        for col_idx, col_name in enumerate(df.columns, start=2):
            if not pd.api.types.is_numeric_dtype(df[col_name]):
                continue
            if col_name in _rs_count_cols:
                continue  # keep General
            elif "_WoW" in col_name:
                fmt = '0.00%'
            else:
                fmt = '0%'
            for row_idx in range(data_start_row, data_start_row + len(df)):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = fmt
        return

    for col_idx, col_name in enumerate(df.columns, start=2):
        if col_name in NON_PCT_COLS:
            continue
        if not pd.api.types.is_numeric_dtype(df[col_name]):
            continue
        should_fmt_pct = False
        if is_pct_sheet(sheet_name) and col_name not in NON_PCT_COLS:
            should_fmt_pct = True
        elif col_name.endswith("_%") or col_name.endswith("_pct"):
            should_fmt_pct = True
        elif "Part%" in col_name or "GotMsg%" in col_name:
            should_fmt_pct = True
        if should_fmt_pct:
            for row_idx in range(data_start_row, data_start_row + len(df)):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = '0.0%'


# ─── Group-header definitions for merged top rows ────────────────────────────
# Maps a column-name prefix to the display label for the merged group header.
# Each metric has 3 sub-columns (Snapp, Jnt_Snapp, Jnt_Tapsi).
SAT_GROUP_HEADERS = [
    ("Part%",        "% Incentive Participation"),
    ("Part_GotMsg%", "% Incentive Participation (Got Message)"),
    ("Incentive_Sat", "Avg Incentive Satisfaction (1-5)"),
    ("Fare_Sat",     "Avg Fare Satisfaction (1-5)"),
    ("Request_Sat",  "Avg Request Satisfaction (1-5)"),
    ("Income_Sat",   "Avg Income Satisfaction (1-5)"),
]

DURATION_GROUP_HEADERS = [
    ("Snapp_", "Snapp Incentive Duration"),
    ("Tapsi_", "Tapsi Incentive Duration"),
]

# Sub-columns under each incentive-type group
INCTYPE_SUBCOLS = ["Exclusives Snapp", "Joints Snapp", "Tapsi"]

# Populated at runtime by analysis_received_incentive_types()
INCTYPE_GROUP_HEADERS = []

# Populated at runtime by analysis_incentive_dissatisfaction()
DISSAT_GROUP_HEADERS = []

# Populated at runtime by analysis_all_cities_overview()
CITIES_GROUP_HEADERS = []

# Populated at runtime by analysis_ride_share()
RIDESHARE_GROUP_HEADERS = []

# Populated at runtime by analysis_driver_persona_parttime_rides()
PERSONA_PARTTIME_GROUP_HEADERS = []

# Static group headers for #18 Commission Free sheets
COMMFREE_GROUP_HEADERS = [
    ("GotMsg_", "Got Message:"),
    ("Participation_", "%Participation Among Who Got:"),
    ("RidesAmong_", "Among Who had Free Commission Ride"),
    ("FreeCommShare_", "Free Commission Ride Share:"),
]

# Map sheet prefix → group header definitions (for merged top row)
SHEET_GROUP_HEADERS = {
    "#3_Sat_": SAT_GROUP_HEADERS,
    "#4_Incentive_Duration": DURATION_GROUP_HEADERS,
    "#5_6_IncType": INCTYPE_GROUP_HEADERS,
    "#8_Dissat": DISSAT_GROUP_HEADERS,
    "#12_Cities_Overview": CITIES_GROUP_HEADERS,
    "#13_RideShare": RIDESHARE_GROUP_HEADERS,
    "#15_Persona_PartTime": PERSONA_PARTTIME_GROUP_HEADERS,
    "#18_CommFree_": COMMFREE_GROUP_HEADERS,
}


def add_group_header_row(ws, df, thin_border, group_headers=None):
    """Insert a merged group-header row (row 1) above column headers.

    Shifts existing content down by 1 row first.
    """
    ws.insert_rows(1)

    group_fill = PatternFill(start_color="4472C4",
                             end_color="4472C4", fill_type="solid")
    group_font = Font(bold=True, color="FFFFFF", size=11)
    group_align = Alignment(horizontal="center", vertical="center")

    # Build column-index ranges for each group
    if group_headers is None:
        group_headers = SAT_GROUP_HEADERS
    for prefix, label in group_headers:
        # Find all columns belonging to this group (row 2 has col headers now)
        col_indices = []
        for col_idx, col_name in enumerate(df.columns, start=2):
            if col_name.startswith(prefix):
                col_indices.append(col_idx)
        if not col_indices:
            continue
        start_col = min(col_indices)
        end_col = max(col_indices)
        # Merge & write label
        if start_col < end_col:
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col, value=label)
        cell.font = group_font
        cell.fill = group_fill
        cell.alignment = group_align
        cell.border = thin_border

    # Style the empty cells in row 1 (n, n_joint, City label)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.border = thin_border


def crosstab_by_city(df, col, top_cities=True):
    """Standard crosstab helper: % distribution of col by city, with n."""
    if top_cities:
        df = filter_top_cities(df)
    valid = df[df[col].notna()]
    if valid.empty:
        return pd.DataFrame()
    ct = pd.crosstab(valid["city"], valid[col], normalize="index") * 100
    ct = ct.round(1)
    ct["n"] = valid.groupby("city").size()
    ct = add_total_row(ct, len(valid))
    ct.index.name = "City"
    return ct


def mean_by_city(df, cols, top_cities=True):
    """Standard mean-by-city helper for satisfaction-type cols."""
    if top_cities:
        df = filter_top_cities(df)
    available = [c for c in cols if c in df.columns]
    if not available:
        return pd.DataFrame()
    for c in available:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    means = df.groupby("city")[available].mean().round(2)
    means["n"] = df.groupby("city")[available[0]].apply(
        lambda x: x.notna().sum())
    means = add_total_row(means)
    means.index.name = "City"
    return means


# ═══════════════════════════════════════════════════════════════════════════
#  ANALYSIS FUNCTIONS — from short_main / wide_main / long_main
# ═══════════════════════════════════════════════════════════════════════════

def analysis_incentive_amounts_snapp(data, week):
    print("\n[#1] Snapp Incentive Amounts Distribution...")
    df = filter_top_cities(filter_week(data["short_main"], week))
    return crosstab_by_city(df, "snapp_incentive_rial_details", top_cities=False)


def analysis_incentive_amounts_tapsi(data, week):
    print("\n[#2] Tapsi Incentive Amounts (Joint Drivers)...")
    df = filter_top_cities(filter_week(data["short_main"], week))
    df = df[df["active_joint"] == 1]
    return crosstab_by_city(df, "tapsi_incentive_rial_details", top_cities=False)


def analysis_satisfaction_review(data, week):
    """Satisfaction & Participation Review — 3 sheets (All / Part-Time / Full-Time).

    Each sheet has cities as rows.  For every metric we compute three driver
    segments side-by-side:
        All Snapp  |  Joint@Snapp (active_joint==1 & snapp_participation==Yes)
                   |  Joint@Tapsi (active_joint==1 & tapsi_participation==Yes)

    Metrics (6 groups):
      1. % Incentive Participation
      2. % Incentive Participation (Among Who Got Messages)
      3. Avg Incentive Satisfaction (1-5)
      4. Avg Fare Satisfaction (1-5)
      5. Avg Request Satisfaction (1-5)
      6. Avg Income Satisfaction (1-5)
    """
    print("\n[#3] Satisfaction / Participation Review...")
    sm = data["short_main"]

    # ── helpers ──────────────────────────────────────────────────────────
    def _pct_yes(series):
        """% of 'Yes' among ALL values (NaN counts as non-Yes)."""
        return (series == "Yes").sum() / len(series) * 100 if len(series) > 0 else np.nan

    def _participation_rate(grp, col):
        return grp[col].apply(_pct_yes).round(1)

    def _sat_mean(grp, col):
        return grp[col].apply(
            lambda x: pd.to_numeric(x, errors="coerce").mean()
        ).round(2)

    # Metric definitions: (display_name, compute_fn)
    # compute_fn(df_segment, city_groups) → Series indexed by city
    def _build_metrics(df):
        """Return list of (metric_label, snapp_series, tapsi_series)."""
        g_all = df.groupby("city")
        g_joint = df[df["active_joint"] == 1].groupby("city")

        metrics = []

        # 1) % Incentive Participation
        #    AllSnapp denom = all respondents; Joint denom = all joint drivers
        snapp_part = _participation_rate(
            g_all, "snapp_incentive_participation")
        snapp_part_joint = _participation_rate(
            g_joint, "snapp_incentive_participation")
        tapsi_part = _participation_rate(
            g_joint, "tapsi_incentive_participation")
        metrics.append(("Part%", snapp_part, snapp_part_joint, tapsi_part))

        # 2) % Incentive Participation (Among Who Got Messages)
        got_msg_sn = df[df["snapp_gotmessage_text_incentive"] == "Yes"]
        got_msg_tp = df[(df["active_joint"] == 1) &
                        (df["tapsi_gotmessage_text_incentive"] == "Yes")]
        got_msg_joint = df[(df["active_joint"] == 1) &
                           (df["snapp_gotmessage_text_incentive"] == "Yes")]
        sn_msg = _participation_rate(
            got_msg_sn.groupby("city"), "snapp_incentive_participation")
        sn_msg_joint = _participation_rate(
            got_msg_joint.groupby("city"), "snapp_incentive_participation")
        tp_msg = _participation_rate(
            got_msg_tp.groupby("city"), "tapsi_incentive_participation")
        metrics.append(("Part_GotMsg%", sn_msg, sn_msg_joint, tp_msg))

        # 3-6) Satisfaction averages
        sat_map = [
            ("Incentive_Sat", "snapp_overall_incentive_satisfaction",
             "tapsi_overall_incentive_satisfaction"),
            ("Fare_Sat", "snapp_fare_satisfaction",
             "tapsi_fare_satisfaction"),
            ("Request_Sat", "snapp_req_count_satisfaction",
             "tapsi_req_count_satisfaction"),
            ("Income_Sat", "snapp_income_satisfaction",
             "tapsi_income_satisfaction"),
        ]
        for label, sn_col, tp_col in sat_map:
            sn_all = _sat_mean(
                g_all, sn_col) if sn_col in df.columns else pd.Series(dtype=float)
            sn_joint = _sat_mean(
                g_joint, sn_col) if sn_col in df.columns else pd.Series(dtype=float)
            tp_joint = _sat_mean(
                g_joint, tp_col) if tp_col in df.columns else pd.Series(dtype=float)
            metrics.append((label, sn_all, sn_joint, tp_joint))

        return metrics

    def _build_sheet(curr_df):
        """Build a single combined DataFrame for one driver segment."""
        curr_metrics = _build_metrics(curr_df)

        combined = pd.DataFrame(index=sorted(
            curr_df["city"].dropna().unique()))

        # Sample counts
        combined["n"] = curr_df.groupby("city").size()
        combined["n_joint"] = curr_df[curr_df["active_joint"]
                                      == 1].groupby("city").size()

        for label, sn_all, sn_joint, tp_joint in curr_metrics:
            combined[f"{label}_Snapp"] = sn_all
            combined[f"{label}_Jnt_Snapp"] = sn_joint
            combined[f"{label}_Jnt_Tapsi"] = tp_joint

        # Total row
        total = pd.DataFrame({"n": [len(curr_df)]}, index=["Total"])
        total["n_joint"] = (curr_df["active_joint"] == 1).sum()
        for c in combined.columns:
            if c in ("n", "n_joint"):
                continue
            total[c] = combined[c].mean()
        combined = pd.concat([combined, total]).round(2)
        combined.index.name = "City"
        return sort_cities(combined)

    # ── Build 3 sheets ──────────────────────────────────────────────────
    results = {}
    for seg_name, seg_fn in [
        ("All Drivers", lambda d: d),
        ("Part-Time", lambda d: d[d["cooperation_type"] == "Part-Time"]),
        ("Full-Time", lambda d: d[d["cooperation_type"] == "Full-Time"]),
    ]:
        curr = seg_fn(filter_top_cities(filter_week(sm, week)))
        if curr.empty:
            continue
        results[seg_name] = _build_sheet(curr)

    return results


def analysis_incentive_time_limitation(data, week):
    print("\n[#4] Incentive Time Limitation...")
    df = filter_top_cities(filter_week(data["short_main"], week))
    combined = pd.DataFrame()
    for platform, col in [
        ("Snapp", "snapp_incentive_active_duration"),
        ("Tapsi", "tapsi_incentive_active_duration"),
    ]:
        if col not in df.columns:
            continue
        ct = crosstab_by_city(df, col, top_cities=False)
        if ct.empty:
            continue
        # Prefix all columns with platform name
        ct = ct.rename(columns={c: f"{platform}_{c}" for c in ct.columns})
        if combined.empty:
            combined = ct
        else:
            combined = combined.join(ct, how="outer")
    return combined


def analysis_received_incentive_types(data, week):
    """Single combined sheet: each incentive-type answer becomes a group header,
    with sub-columns Exclusives Snapp / Joints Snapp / Tapsi.

    Formula matches the reference Excel:
      Denominator per segment:
        Exclusives Snapp = exclusive drivers (active_joint=0) who got Snapp msg
        Joints Snapp     = joint drivers (active_joint=1) who got Snapp msg
        Tapsi            = joint drivers (active_joint=1) who got Tapsi msg
      Numerator = same filter + specific incentive type binary = 1
    """
    print("\n[#5/#6] Received Incentive Types...")
    sm = filter_top_cities(filter_week(data["short_main"], week))
    wm = filter_top_cities(filter_week(data["wide_main"], week))

    # Identify incentive-type binary columns in wide_main
    snapp_type_cols = {c: c.replace("Snapp Incentive Type__", "")
                       for c in wm.columns if c.startswith("Snapp Incentive Type__")}
    tapsi_type_cols = {c: c.replace("Tapsi Incentive Type__", "")
                       for c in wm.columns if c.startswith("Tapsi Incentive Type__")}

    # Merge needed short_main columns into wide_main
    merge_cols = ["recordID", "city", "active_joint",
                  "snapp_gotmessage_text_incentive",
                  "tapsi_gotmessage_text_incentive"]
    merge_cols = [c for c in merge_cols if c in sm.columns]
    wm = wm.merge(sm[merge_cols].drop_duplicates(), on="recordID", how="left",
                  suffixes=("_wm", ""))

    # Three segments with their denominators
    ex_snapp = wm[(wm["active_joint"] == 0) &
                  (wm["snapp_gotmessage_text_incentive"] == "Yes")]
    jnt_snapp = wm[(wm["active_joint"] == 1) &
                   (wm["snapp_gotmessage_text_incentive"] == "Yes")]
    jnt_tapsi = wm[(wm["active_joint"] == 1) &
                   (wm["tapsi_gotmessage_text_incentive"] == "Yes")]

    def _pct_by_city(seg_df, type_cols_map):
        """Compute % per city: for each type, count(type=1) / count(segment)."""
        if seg_df.empty:
            return pd.DataFrame(), pd.Series(dtype="int64")
        denom = seg_df.groupby("city").size()
        result = pd.DataFrame(index=denom.index)
        for raw_col, answer_label in type_cols_map.items():
            if raw_col in seg_df.columns:
                nums = seg_df.groupby("city")[raw_col].sum()
                result[answer_label] = (nums / denom * 100).round(1)
        result.index.name = "City"
        return result, denom

    pct_ex, n_ex = _pct_by_city(ex_snapp, snapp_type_cols)
    pct_jnt, n_jnt = _pct_by_city(jnt_snapp, snapp_type_cols)
    pct_tapsi, n_tapsi = _pct_by_city(jnt_tapsi, tapsi_type_cols)

    # Collect all unique answer labels
    all_answers = sorted(set(
        list(pct_ex.columns if not pct_ex.empty else [])
        + list(pct_jnt.columns if not pct_jnt.empty else [])
        + list(pct_tapsi.columns if not pct_tapsi.empty else [])
    ))

    # Build group headers and combined DataFrame
    INCTYPE_GROUP_HEADERS.clear()
    combined = pd.DataFrame()
    col_order = []

    for answer in all_answers:
        prefix = answer.replace(" ", "_").replace("-", "_") + "_"
        INCTYPE_GROUP_HEADERS.append((prefix, answer))

        for sub, pct_df in [("Exclusives Snapp", pct_ex),
                            ("Joints Snapp", pct_jnt),
                            ("Tapsi", pct_tapsi)]:
            col_name = f"{prefix}{sub}"
            col_order.append(col_name)
            if not pct_df.empty and answer in pct_df.columns:
                combined[col_name] = pct_df[answer]
            else:
                combined[col_name] = pd.NA

    # Count columns: n for each segment's denominator
    combined["ExSnapp_n"] = n_ex if not n_ex.empty else pd.NA
    combined["JntSnapp_n"] = n_jnt if not n_jnt.empty else pd.NA
    combined["Tapsi_n"] = n_tapsi if not n_tapsi.empty else pd.NA
    col_order += ["ExSnapp_n", "JntSnapp_n", "Tapsi_n"]

    combined.index.name = "City"

    # Add Total row
    sm_all = filter_week(data["short_main"], week)
    wm_all = filter_week(data["wide_main"], week)
    wm_all = wm_all.merge(
        sm_all[merge_cols].drop_duplicates(), on="recordID", how="left",
        suffixes=("_wm", ""))
    ex_all = wm_all[(wm_all["active_joint"] == 0) &
                    (wm_all["snapp_gotmessage_text_incentive"] == "Yes")]
    jnt_sn_all = wm_all[(wm_all["active_joint"] == 1) &
                        (wm_all["snapp_gotmessage_text_incentive"] == "Yes")]
    jnt_tp_all = wm_all[(wm_all["active_joint"] == 1) &
                        (wm_all["tapsi_gotmessage_text_incentive"] == "Yes")]

    total_data = {}
    for answer in all_answers:
        prefix = answer.replace(" ", "_").replace("-", "_") + "_"
        for sub, seg_df, type_cols in [
            ("Exclusives Snapp", ex_all, snapp_type_cols),
            ("Joints Snapp", jnt_sn_all, snapp_type_cols),
            ("Tapsi", jnt_tp_all, tapsi_type_cols),
        ]:
            col_name = f"{prefix}{sub}"
            raw_col = next(
                (k for k, v in type_cols.items() if v == answer), None)
            if raw_col and raw_col in seg_df.columns and len(seg_df) > 0:
                total_data[col_name] = round(
                    seg_df[raw_col].sum() / len(seg_df) * 100, 1)
            else:
                total_data[col_name] = pd.NA
    total_data["ExSnapp_n"] = len(ex_all)
    total_data["JntSnapp_n"] = len(jnt_sn_all)
    total_data["Tapsi_n"] = len(jnt_tp_all)
    combined.loc["Total"] = total_data

    # Reorder columns
    combined = combined[[c for c in col_order if c in combined.columns]]
    return combined


def analysis_incentive_dissatisfaction(data, week):
    """Combined sheet: All Snapp dissatisfaction | Joint Tapsi dissatisfaction.

    Formula matches the reference Excel:
      All Snapp:
        Denominator = ALL drivers in city where snapp_overall_incentive_satisfaction < 4
        Numerator   = same filter + specific reason binary col = 1
        Uses "Snapp Last Incentive Unsatisfaction__*" wide_main columns
      Joint Tapsi:
        Denominator = JOINT drivers (active_joint=1) where tapsi_overall_incentive_satisfaction < 4
        Numerator   = same filter + specific Tapsi reason binary col = 1
        Uses "Tapsi Incentive Unsatisfaction__*" wide_main columns
    """
    print("\n[#8/#9] Incentive Dissatisfaction Reasons...")
    sm = filter_top_cities(filter_week(data["short_main"], week))
    wm = filter_top_cities(filter_week(data["wide_main"], week))

    # Merge needed short_main columns
    merge_cols = ["recordID", "city", "active_joint",
                  "snapp_overall_incentive_satisfaction",
                  "tapsi_overall_incentive_satisfaction"]
    merge_cols = [c for c in merge_cols if c in sm.columns]
    wm = wm.merge(sm[merge_cols].drop_duplicates(), on="recordID", how="left",
                  suffixes=("_wm", ""))

    # Identify binary reason columns
    snapp_reason_cols = {c: c.replace("Snapp Last Incentive Unsatisfaction__", "")
                         for c in wm.columns
                         if c.startswith("Snapp Last Incentive Unsatisfaction__")}
    tapsi_reason_cols = {c: c.replace("Tapsi Incentive Unsatisfaction__", "")
                         for c in wm.columns
                         if c.startswith("Tapsi Incentive Unsatisfaction__")}

    # --- All Snapp: all drivers with incentive_sat <= 3 ---
    snapp_dissat = wm[wm["snapp_overall_incentive_satisfaction"].le(3)]
    # --- Joint Tapsi: joint drivers with incentive_sat <= 3 ---
    tapsi_dissat = wm[(wm["active_joint"] == 1) &
                      wm["tapsi_overall_incentive_satisfaction"].le(3)]

    def _pct_by_city(seg_df, reason_cols_map):
        if seg_df.empty:
            return pd.DataFrame(), pd.Series(dtype="int64")
        denom = seg_df.groupby("city").size()
        result = pd.DataFrame(index=denom.index)
        for raw_col, answer_label in reason_cols_map.items():
            if raw_col in seg_df.columns:
                nums = seg_df.groupby("city")[raw_col].sum()
                result[answer_label] = (nums / denom * 100).round(1)
        result.index.name = "City"
        return result, denom

    pct_snapp, n_snapp = _pct_by_city(snapp_dissat, snapp_reason_cols)
    pct_tapsi, n_tapsi = _pct_by_city(tapsi_dissat, tapsi_reason_cols)

    # Collect all unique answer labels
    all_answers = sorted(set(
        list(pct_snapp.columns if not pct_snapp.empty else [])
        + list(pct_tapsi.columns if not pct_tapsi.empty else [])
    ))

    # Build group headers and combined DataFrame
    DISSAT_GROUP_HEADERS.clear()
    DISSAT_GROUP_HEADERS.append(("Snapp_", "All Snapp"))
    DISSAT_GROUP_HEADERS.append(("Tapsi_", "Joint Tapsi"))

    combined = pd.DataFrame()
    col_order = []

    # Snapp columns
    for answer in all_answers:
        col_name = f"Snapp_{answer}"
        col_order.append(col_name)
        if not pct_snapp.empty and answer in pct_snapp.columns:
            combined[col_name] = pct_snapp[answer]
        else:
            combined[col_name] = pd.NA
    combined["Snapp_n"] = n_snapp if not n_snapp.empty else pd.NA
    col_order.append("Snapp_n")

    # Tapsi columns
    for answer in all_answers:
        col_name = f"Tapsi_{answer}"
        col_order.append(col_name)
        if not pct_tapsi.empty and answer in pct_tapsi.columns:
            combined[col_name] = pct_tapsi[answer]
        else:
            combined[col_name] = pd.NA
    combined["Tapsi_n"] = n_tapsi if not n_tapsi.empty else pd.NA
    col_order.append("Tapsi_n")

    combined.index.name = "City"

    # Add Total row (unfiltered by top cities)
    sm_all = filter_week(data["short_main"], week)
    wm_all = filter_week(data["wide_main"], week)
    wm_all = wm_all.merge(
        sm_all[merge_cols].drop_duplicates(), on="recordID", how="left",
        suffixes=("_wm", ""))
    sn_all = wm_all[wm_all["snapp_overall_incentive_satisfaction"].le(3)]
    tp_all = wm_all[(wm_all["active_joint"] == 1) &
                    wm_all["tapsi_overall_incentive_satisfaction"].le(3)]
    total_data = {}
    for answer in all_answers:
        for prefix, seg_df, reason_cols in [
            ("Snapp_", sn_all, snapp_reason_cols),
            ("Tapsi_", tp_all, tapsi_reason_cols),
        ]:
            col_name = f"{prefix}{answer}"
            raw_col = next(
                (k for k, v in reason_cols.items() if v == answer), None)
            if raw_col and raw_col in seg_df.columns and len(seg_df) > 0:
                total_data[col_name] = round(
                    seg_df[raw_col].sum() / len(seg_df) * 100, 1)
            else:
                total_data[col_name] = pd.NA
    total_data["Snapp_n"] = len(sn_all)
    total_data["Tapsi_n"] = len(tp_all)
    combined.loc["Total"] = total_data

    # Reorder columns
    combined = combined[[c for c in col_order if c in combined.columns]]

    # --- Summary (#9): reasons as rows, segments as columns ---
    # Three segments: All Snapp, Joint Snapp (active_joint=1 + snapp sat<=3),
    #                 Joint Tapsi (active_joint=1 + tapsi sat<=3)
    # Display answer names matching reference Excel
    ANSWER_DISPLAY = {
        "Not Available": "I Have no Enough Time",
        "Improper Amount": "Improper Incentive Amount",
        "No Time todo": "Low Time",
        "difficult": "Hard to Do",
        "Non Payment": "Non-Payment",
    }
    REASON_ORDER = [
        "I Have no Enough Time", "Improper Incentive Amount",
        "Low Time", "Hard to Do", "Non-Payment",
    ]

    def _build_dissat_summary(wm_week):
        """Build reason-pct Series for each of the 3 segments."""
        sn = wm_week[wm_week["snapp_overall_incentive_satisfaction"].le(3)]
        jnt_sn = wm_week[(wm_week["active_joint"] == 1) &
                          wm_week["snapp_overall_incentive_satisfaction"].le(3)]
        jnt_tp = wm_week[(wm_week["active_joint"] == 1) &
                          wm_week["tapsi_overall_incentive_satisfaction"].le(3)]
        segs = {}
        for seg_label, seg_df, rcols in [
            ("All Snapp", sn, snapp_reason_cols),
            ("Joint Snapp", jnt_sn, snapp_reason_cols),
            ("Joint Tapsi", jnt_tp, tapsi_reason_cols),
        ]:
            d = len(seg_df)
            pcts = {}
            for raw_col, ans in rcols.items():
                display = ANSWER_DISPLAY.get(ans, ans)
                if raw_col in seg_df.columns and d > 0:
                    pcts[display] = round(seg_df[raw_col].sum() / d * 100, 1)
            segs[seg_label] = pd.Series(pcts)
            segs[f"{seg_label}_n"] = d
        return segs

    # Current week
    curr_segs = _build_dissat_summary(wm_all)
    # Previous week
    prev_week = week - 1
    sm_prev = filter_week(data["short_main"], prev_week)
    wm_prev = filter_week(data["wide_main"], prev_week)
    if not sm_prev.empty and not wm_prev.empty:
        prev_merge = [c for c in merge_cols if c in sm_prev.columns]
        wm_prev = wm_prev.merge(
            sm_prev[prev_merge].drop_duplicates(), on="recordID",
            how="left", suffixes=("_wm", ""))
        prev_segs = _build_dissat_summary(wm_prev)
    else:
        prev_segs = {}

    # Build summary DataFrame: rows=reasons, cols=segments+WoW
    summary_df = pd.DataFrame(index=REASON_ORDER)
    summary_df.index.name = "Reasons"

    for seg in ["All Snapp", "Joint Snapp", "Joint Tapsi"]:
        curr_s = curr_segs.get(seg, pd.Series(dtype=float))
        prev_s = prev_segs.get(seg, pd.Series(dtype=float))
        summary_df[seg] = curr_s.reindex(REASON_ORDER)
        if not prev_s.empty:
            wow = (curr_s.reindex(REASON_ORDER) - prev_s.reindex(REASON_ORDER)).round(1)
            summary_df[f"{seg}_WoW"] = wow
        else:
            summary_df[f"{seg}_WoW"] = pd.NA

    # Build last-week DataFrame
    last_week_df = pd.DataFrame(index=REASON_ORDER)
    last_week_df.index.name = "Reasons"
    for seg in ["All Snapp", "Joint Snapp", "Joint Tapsi"]:
        prev_s = prev_segs.get(seg, pd.Series(dtype=float))
        last_week_df[seg] = prev_s.reindex(REASON_ORDER) if not prev_s.empty else pd.NA
        # WoW for last week would need week-2 data; leave blank or omit
        # Include previous WoW column from the prev week's reference (week-2)
        # For simplicity just include the values without WoW

    # Reorder columns: seg, wow, seg, wow, seg, wow
    col_order_summary = []
    for seg in ["All Snapp", "Joint Snapp", "Joint Tapsi"]:
        col_order_summary.append(seg)
        col_order_summary.append(f"{seg}_WoW")
    summary_df = summary_df[[c for c in col_order_summary if c in summary_df.columns]]

    return {
        "combined": combined,
        "summary": summary_df,
        "summary_last_week": last_week_df,
    }


def analysis_all_cities_overview(data, week):
    """Build #12 Cities Overview matching reference Excel format.

    Columns:
        E_n (# Respondents), F_n (# Joint), G_n (# Cmpt Signup),
        % Joint (H), % Dual SU (I),
        Avg LOC: All Snapp (J), Joint Snapp (K), Joint Cmpt (L), Joint Cmpt SU (M),
        Got Msg %: All Snapp (N), Joint Snapp (O), Joint Cmpt (P), Joint Cmpt SU (Q)

    Three independent cutoffs from SHEET_MIN_N (default 17 each):
        E_n > cutoff  → % Joint, % Dual SU, AvgLOC_All Snapp, GotMsg_All Snapp
        F_n > cutoff  → AvgLOC_Joint Snapp, GotMsg_Joint Snapp, GotMsg_Joint Cmpt
        G_n > cutoff  → AvgLOC_Joint Cmpt, AvgLOC_Joint Cmpt SU, GotMsg_Joint Cmpt SU
    """
    print("\n[#12] All Cities Overview...")
    df = filter_top_cities(filter_week(data["short_main"], week))
    if df.empty:
        return pd.DataFrame()

    rows = []
    for city, grp in df.groupby("city"):
        E = len(grp)
        F = int((grp["active_joint"] == 1).sum())
        G = int((grp["tapsi_LOC"] > 0).sum())
        joint = grp[grp["active_joint"] == 1]
        cmpt_su = grp[grp["tapsi_LOC"] > 0]

        row = {"E_n": E, "F_n": F, "G_n": G}

        # --- E group: all drivers ---
        row["% Joint"] = F / E * 100 if E else np.nan
        row["% Dual SU"] = G / E * 100 if E else np.nan
        row["AvgLOC_All Snapp"] = grp["snapp_LOC"].mean()
        snapp_msg_yes = (grp["snapp_gotmessage_text_incentive"] == "Yes").sum()
        row["GotMsg_All Snapp"] = snapp_msg_yes / E * 100 if E else np.nan

        # --- F group: joint drivers ---
        row["AvgLOC_Joint Snapp"] = joint["snapp_LOC"].mean() if F else np.nan
        jnt_snapp_msg = (joint["snapp_gotmessage_text_incentive"] == "Yes").sum()
        row["GotMsg_Joint Snapp"] = jnt_snapp_msg / F * 100 if F else np.nan
        jnt_tapsi_msg = (joint["tapsi_gotmessage_text_incentive"] == "Yes").sum()
        row["GotMsg_Joint Cmpt"] = jnt_tapsi_msg / F * 100 if F else np.nan

        # --- G group: cmpt signup drivers ---
        row["AvgLOC_Joint Cmpt"] = joint["tapsi_LOC"].mean() if F else np.nan
        row["AvgLOC_Joint Cmpt SU"] = cmpt_su["tapsi_LOC"].mean() if G else np.nan
        cmpt_su_tapsi_msg = (cmpt_su["tapsi_gotmessage_text_incentive"] == "Yes").sum()
        row["GotMsg_Joint Cmpt SU"] = cmpt_su_tapsi_msg / G * 100 if G else np.nan

        rows.append((city, row))

    # Build DataFrame
    col_order = [
        "E_n", "F_n", "G_n",
        "% Joint", "% Dual SU",
        "AvgLOC_All Snapp", "AvgLOC_Joint Snapp",
        "AvgLOC_Joint Cmpt", "AvgLOC_Joint Cmpt SU",
        "GotMsg_All Snapp", "GotMsg_Joint Snapp",
        "GotMsg_Joint Cmpt", "GotMsg_Joint Cmpt SU",
    ]
    result = pd.DataFrame([r for _, r in rows],
                          index=[c for c, _ in rows],
                          columns=col_order)
    result.index.name = "City"

    # Reorder to TOP_CITIES order
    result = sort_cities(result)

    # Populate group headers for merged top row
    CITIES_GROUP_HEADERS.clear()
    CITIES_GROUP_HEADERS.append(("AvgLOC_", "Avg LOC (Month)"))
    CITIES_GROUP_HEADERS.append(("GotMsg_", "% Got Incentive Messages"))

    return result


def analysis_ride_share(data, week):
    """Build #13 Ride Share — single combined sheet matching reference Excel.

    Left side (counts):
        total Res, Joint Res, Ex drivers,
        Total Ride (sum snapp+tapsi), Total Ride Snapp, Ex drivers Ride in Snapp,
        Snapp Ride (joint), Tapsi Tide (joint)
    Right side (percentages + WoW):
        All Snapp % (=TotalRideSnapp/TotalRide), Last, WoW,
        Ex Drivers in Snapp % (=ExRideSnapp/TotalRide), Last, WoW,
        Joint @Snapp % (=JntSnappRide/TotalRide), Last, WoW,
        Joint @Tapsi % (=JntTapsiRide/TotalRide), Last, WoW

    Min-N cutoff: F_n (total Res) >= threshold from SHEET_MIN_N.
    Percentages left blank for cities below cutoff.
    """
    print("\n[#13] Drivers' Ride Share...")
    MIN_N = 1  # reference uses D1=1

    def _build_ride_share(sm_week):
        df = filter_top_cities(sm_week)
        if df.empty:
            return pd.DataFrame()
        rows = []
        for city, grp in df.groupby("city"):
            F = len(grp)
            G = int((grp["active_joint"] == 1).sum())
            H = int((grp["active_joint"] == 0).sum())
            joint = grp[grp["active_joint"] == 1]
            excl = grp[grp["active_joint"] == 0]

            snapp_total = grp["snapp_ride"].sum()
            tapsi_total = grp["tapsi_ride"].sum()
            I = snapp_total + tapsi_total  # Total Ride
            J = snapp_total                # Total Ride Snapp
            K = excl["snapp_ride"].sum()   # Ex drivers Ride in Snapp
            L = joint["snapp_ride"].sum()  # Snapp Ride (joint)
            M = joint["tapsi_ride"].sum()  # Tapsi Tide (joint)

            row = {
                "total Res": F, "Joint Res": G, "Ex drivers": H,
                "Total Ride": I, "Total Ride Snapp": J,
                "Ex drivers Ride in Snapp": K,
                "Snapp Ride": L, "Tapsi Tide": M,
            }
            # Percentages: ride share ratios (stored as 0-100, converted to decimal later)
            if F >= MIN_N and I > 0:
                row["All Snapp"] = J / I * 100
            if H >= MIN_N and I > 0:
                row["Ex Drivers in Snapp"] = K / I * 100
            if G >= MIN_N and I > 0:
                row["Jnt @Snapp"] = L / I * 100
                row["Jnt @Tapsi"] = M / I * 100

            rows.append((city, row))
        result = pd.DataFrame([r for _, r in rows],
                              index=[c for c, _ in rows])
        result.index.name = "City"
        return result

    curr_df = _build_ride_share(filter_week(data["short_main"], week))
    if curr_df.empty:
        return pd.DataFrame()

    # Previous week for WoW
    prev_df = _build_ride_share(filter_week(data["short_main"], week - 1))

    pct_cols = ["All Snapp", "Ex Drivers in Snapp", "Jnt @Snapp", "Jnt @Tapsi"]
    count_cols = ["total Res", "Joint Res", "Ex drivers",
                  "Total Ride", "Total Ride Snapp", "Ex drivers Ride in Snapp",
                  "Snapp Ride", "Tapsi Tide"]

    # Build final DataFrame with interleaved Last + WoW columns
    final = curr_df[count_cols].copy()
    for pc in pct_cols:
        final[pc] = curr_df[pc] if pc in curr_df.columns else np.nan
        if not prev_df.empty and pc in prev_df.columns:
            last_col = prev_df[pc].reindex(final.index)
            final[f"{pc}_Last"] = last_col
            final[f"{pc}_WoW"] = final[pc] - last_col
        else:
            final[f"{pc}_Last"] = np.nan
            final[f"{pc}_WoW"] = np.nan

    final = sort_cities(final)

    # Populate group headers
    RIDESHARE_GROUP_HEADERS.clear()
    RIDESHARE_GROUP_HEADERS.append(("All Snapp", "All Snapp"))
    RIDESHARE_GROUP_HEADERS.append(("Ex Drivers in Snapp", "Ex Drivers in Snapp"))
    RIDESHARE_GROUP_HEADERS.append(("Jnt ", "Joint Drivers"))

    return final


def analysis_navigation_usage(data, week):
    print("\n[#14] Navigation App Usage...")
    df = filter_top_cities(filter_week(data["short_main"], week))
    results = {}
    for col, label in [
        ("tapsi_navigation_type", "Tapsi Nav Type"),
        ("snapp_last_trip_navigation", "Snapp Last Trip Nav"),
    ]:
        valid = df[df[col].notna()]
        if valid.empty:
            continue
        ct = pd.crosstab(valid["city"], valid[col], normalize="index") * 100
        ct = ct.round(1)
        ct["n"] = valid.groupby("city").size()
        ct.index.name = "City"
        results[label] = ct

    lr = filter_top_cities(add_city(filter_week(
        data["long_rare"], week), data["_lookup"]))
    for q in ["Navigation Familiarity", "Navigation Installed", "Navigation Used"]:
        qdf = lr[lr["question"] == q]
        if qdf.empty:
            continue
        resp = qdf.groupby("city")["recordID"].nunique()
        ans = qdf.groupby(["city", "answer"]).size().unstack(fill_value=0)
        pct = (ans.div(resp, axis=0) * 100).round(1)
        pct["n"] = resp
        pct.index.name = "City"
        results[q] = pct
    return results


def analysis_driver_persona(data, week):
    print("\n[#15] Drivers' Persona...")
    df = filter_top_cities(filter_week(data["short_main"], week))
    results = {}
    for col, label in [
        ("active_time", "Activity Type"),
        ("age_group", "Age Group"),
        ("edu", "Education"),
        ("marr_stat", "Marital Status"),
        ("original_job", "Original Job"),
        ("gender", "Gender"),
    ]:
        valid = df[df[col].notna()]
        if valid.empty:
            continue
        ct = pd.crosstab(valid["city"], valid[col], normalize="index") * 100
        ct = ct.round(1)
        ct["n"] = valid.groupby("city").size()
        ct = add_total_row(ct, len(valid))
        ct.index.name = "City"
        results[label] = ct
    return results


def analysis_driver_persona_parttime_rides(data, week):
    """Build #15 Part-Time & Ride Per Boarded sheet matching reference Excel.

    Columns:
      total Res, Joint Res, Ex drivers,
      PT_%Part-Time among_Joint, PT_%Part-Time among_WoW,
      PT_%Part-Time among_Exclusive, PT_%Part-Time among_WoW,
      RidePerBoarded_Snapp, RidePerBoarded_Tapsi, AvgAllRides
    """
    print("\n[#15] Drivers' Persona — Part-Time & Ride Per Boarded...")
    MIN_N = 18  # reference B2=18 for part-time display threshold

    def _build_persona(sm_week):
        df = filter_top_cities(sm_week)
        if df.empty:
            return pd.DataFrame()
        rows = []
        for city, grp in df.groupby("city"):
            total = len(grp)
            joint = grp[grp["active_joint"] == 1]
            excl = grp[grp["active_joint"] == 0]
            n_joint = len(joint)
            n_excl = len(excl)

            row = {
                "total Res": total,
                "Joint Res": n_joint,
                "Ex drivers": n_excl,
            }

            # %Part-Time among Joint
            if n_joint >= MIN_N:
                pt_joint = (joint["cooperation_type"] == "Part-Time").sum()
                row["PT_%Part-Time among_Joint"] = pt_joint / n_joint * 100
            # %Part-Time among Exclusive
            if n_excl >= MIN_N:
                pt_excl = (excl["cooperation_type"] == "Part-Time").sum()
                row["PT_%Part-Time among_Exclusive"] = pt_excl / n_excl * 100

            # Ride Per Boarded in
            if n_joint > 0:
                snapp_rides_joint = joint["snapp_ride"].sum()
                tapsi_rides_joint = joint["tapsi_ride"].sum()
                row["RidePerBoarded_Snapp"] = round(snapp_rides_joint / n_joint)
                row["RidePerBoarded_Tapsi"] = round(tapsi_rides_joint / n_joint)
            if total > 0:
                all_rides = grp["snapp_ride"].sum()
                row["AvgAllRides"] = round(all_rides / total)

            rows.append((city, row))

        result = pd.DataFrame([r for _, r in rows], index=[c for c, _ in rows])
        result.index.name = "City"
        return result

    def _add_total(persona_df, sm_week):
        """Add a Total row computed from all top-city respondents."""
        df = filter_top_cities(sm_week)
        if df.empty:
            return persona_df
        total = len(df)
        joint = df[df["active_joint"] == 1]
        excl = df[df["active_joint"] == 0]
        n_j, n_e = len(joint), len(excl)
        row = {"total Res": total, "Joint Res": n_j, "Ex drivers": n_e}
        if n_j >= MIN_N:
            row["PT_%Part-Time among_Joint"] = (
                (joint["cooperation_type"] == "Part-Time").sum() / n_j * 100)
        if n_e >= MIN_N:
            row["PT_%Part-Time among_Exclusive"] = (
                (excl["cooperation_type"] == "Part-Time").sum() / n_e * 100)
        if n_j > 0:
            row["RidePerBoarded_Snapp"] = round(joint["snapp_ride"].sum() / n_j)
            row["RidePerBoarded_Tapsi"] = round(joint["tapsi_ride"].sum() / n_j)
        if total > 0:
            row["AvgAllRides"] = round(df["snapp_ride"].sum() / total)
        total_df = pd.DataFrame([row], index=["Total"])
        total_df.index.name = "City"
        return pd.concat([persona_df, total_df])

    curr = _build_persona(filter_week(data["short_main"], week))
    if curr.empty:
        return pd.DataFrame()
    curr = _add_total(curr, filter_week(data["short_main"], week))

    # Previous week for WoW
    prev_week_df = filter_week(data["short_main"], week - 1)
    prev = _build_persona(prev_week_df)
    if not prev.empty:
        prev = _add_total(prev, prev_week_df)

    pct_cols = ["PT_%Part-Time among_Joint", "PT_%Part-Time among_Exclusive"]

    # Build final with interleaved WoW columns
    final = curr[["total Res", "Joint Res", "Ex drivers"]].copy()
    for pc in pct_cols:
        final[pc] = curr[pc] if pc in curr.columns else np.nan
        # WoW col keeps the segment suffix for uniqueness: Joint_WoW, Exclusive_WoW
        segment = pc.rsplit("_", 1)[1]  # "Joint" or "Exclusive"
        wow_col = f"PT_%Part-Time among_{segment}_WoW"
        if not prev.empty and pc in prev.columns:
            last_val = prev[pc].reindex(final.index)
            final[wow_col] = (final[pc] - last_val).round(0)
        else:
            final[wow_col] = np.nan

    # Ride columns
    for rc in ["RidePerBoarded_Snapp", "RidePerBoarded_Tapsi", "AvgAllRides"]:
        final[rc] = curr[rc] if rc in curr.columns else np.nan

    final = sort_cities(final)

    # Populate group headers
    PERSONA_PARTTIME_GROUP_HEADERS.clear()
    PERSONA_PARTTIME_GROUP_HEADERS.append(
        ("PT_%Part-Time among_", "%Part-Time among"))
    PERSONA_PARTTIME_GROUP_HEADERS.append(
        ("RidePerBoarded_", "Ride Per Boarded in"))

    return final


def analysis_referral_plan(data, week):
    print("\n[#16] Referral Plan...")
    df = filter_top_cities(filter_week(data["short_main"], week))
    results = {}
    for col, label in [
        ("snapp_joining_bonus", "Snapp Joining Bonus"),
        ("tapsi_joining_bonus", "Tapsi Joining Bonus"),
    ]:
        ct = crosstab_by_city(df, col, top_cities=False)
        if not ct.empty:
            results[label] = ct
    return results


def analysis_inactivity_before_incentive(data, week):
    print("\n[#17] Inactivity Before Tapsi Incentive...")
    df = filter_top_cities(filter_week(data["short_main"], week))
    return crosstab_by_city(df, "tapsi_inactive_b4_incentive", top_cities=False)


def analysis_commission_free(data, week):
    """Build #18 Commission-Free Cost sheets matching reference Excel.

    Returns a dict with two DataFrames:
      - "Tapsi": joint drivers only, tapsi incentive/ride columns
      - "Snapp": all drivers, snapp incentive/ride columns
    """
    print("\n[#18] Commission-Free Analysis...")

    INCENTIVE_CATS = ["Money", "Free-Commission", "Money & Free-commission"]
    MIN_N = 8  # reference B1=8

    def _build_commfree(df_week, platform):
        """Build one commission-free table for a given platform."""
        df = filter_top_cities(df_week)
        if df.empty:
            return pd.DataFrame()

        # Column name prefixes per platform
        if platform == "tapsi":
            got_msg_col = "tapsi_gotmessage_text_incentive"
            cat_col = "tapsi_incentive_category"
            part_col = "tapsi_incentive_participation"
            commfree_col = "tapsi_commfree"
            ride_col = "tapsi_ride"
            filter_joint = True
        else:
            got_msg_col = "snapp_gotmessage_text_incentive"
            cat_col = "snapp_incentive_category"
            part_col = "snapp_incentive_participation"
            commfree_col = "snapp_commfree"
            ride_col = "snapp_ride"
            filter_joint = False

        # Ensure numeric
        for c in [commfree_col, ride_col]:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce")

        rows = []
        for city, grp in df.groupby("city"):
            if filter_joint:
                base = grp[grp["active_joint"] == 1]
            else:
                base = grp
            n = len(base)

            got_msg = base[base[got_msg_col] == "Yes"] if got_msg_col in base.columns else pd.DataFrame()
            n_got_msg = len(got_msg)

            # Got Message breakdown by incentive category
            cat_counts = {}
            if cat_col in got_msg.columns and not got_msg.empty:
                vc = got_msg[cat_col].value_counts()
                for cat in INCENTIVE_CATS:
                    cat_counts[cat] = int(vc.get(cat, 0))
            else:
                for cat in INCENTIVE_CATS:
                    cat_counts[cat] = 0

            # Free commission drivers (ALL drivers in city, no joint filter — matching reference)
            if commfree_col in grp.columns:
                commfree_vals = pd.to_numeric(grp[commfree_col], errors="coerce")
                n_free_comm_drivers = int((commfree_vals > 0).sum())
            else:
                n_free_comm_drivers = 0

            row = {"n": n, "Who Got Message": n_got_msg}
            for cat in INCENTIVE_CATS:
                row[f"GotMsg_{cat}"] = cat_counts[cat]
            row["Free Comm Drivers"] = n_free_comm_drivers

            # Percentages (blank if below min-N)
            if n >= MIN_N:
                row["% Who Got Message"] = n_got_msg / n * 100 if n > 0 else np.nan
                # % Free-Commission message = (FreeComm + Money&Free) / got_msg
                if n_got_msg >= MIN_N:
                    fc_msg = cat_counts.get("Free-Commission", 0) + cat_counts.get("Money & Free-commission", 0)
                    row["% Free Comm Message"] = fc_msg / n_got_msg * 100
                # % Free commission ride = drivers with commfree>0 / n
                if filter_joint:
                    # For tapsi: count joint drivers with commfree>0
                    if commfree_col in base.columns:
                        base_cf = pd.to_numeric(base[commfree_col], errors="coerce")
                        n_fc_base = int((base_cf > 0).sum())
                    else:
                        n_fc_base = 0
                    row["% Free Comm Ride"] = n_fc_base / n * 100
                else:
                    row["% Free Comm Ride"] = n_free_comm_drivers / n * 100

            # Participation rates per incentive category
            if part_col in got_msg.columns and not got_msg.empty:
                participated = got_msg[got_msg[part_col] == "Yes"]
                for cat in INCENTIVE_CATS:
                    cat_got = got_msg[got_msg[cat_col] == cat] if cat_col in got_msg.columns else pd.DataFrame()
                    n_cat = len(cat_got)
                    if n_cat >= MIN_N:
                        n_cat_part = len(participated[participated[cat_col] == cat]) if cat_col in participated.columns else 0
                        row[f"Participation_{cat}"] = n_cat_part / n_cat * 100
                # All incentive participation
                if n_got_msg >= MIN_N:
                    row["Participation_All"] = len(participated) / n_got_msg * 100

            # Ride data
            if ride_col in grp.columns:
                ride_vals = pd.to_numeric(grp[ride_col], errors="coerce")
                total_rides = ride_vals.sum()
                row["Total Rides"] = total_rides

                if commfree_col in grp.columns:
                    has_fc = commfree_vals > 0
                    rides_fc = ride_vals[has_fc].sum()
                    fc_rides = commfree_vals[has_fc].sum()
                    row["RidesAmong_Total Rides"] = rides_fc
                    row["RidesAmong_Free Comm Rides"] = fc_rides
                    if n >= MIN_N and total_rides > 0:
                        row["FreeCommShare_Among All"] = fc_rides / total_rides * 100
                    if n >= MIN_N and rides_fc > 0:
                        row["FreeCommShare_Among Free Comm"] = fc_rides / rides_fc * 100

            rows.append((city, row))

        result = pd.DataFrame([r for _, r in rows], index=[c for c, _ in rows])
        result.index.name = "City"

        # Add Total row
        totals = {}
        if filter_joint:
            all_base = df[df["active_joint"] == 1]
        else:
            all_base = df
        n_total = len(all_base)
        got_msg_total = all_base[all_base[got_msg_col] == "Yes"] if got_msg_col in all_base.columns else pd.DataFrame()
        n_got_msg_total = len(got_msg_total)

        totals["n"] = n_total
        totals["Who Got Message"] = n_got_msg_total
        if cat_col in got_msg_total.columns and not got_msg_total.empty:
            vc = got_msg_total[cat_col].value_counts()
            for cat in INCENTIVE_CATS:
                totals[f"GotMsg_{cat}"] = int(vc.get(cat, 0))
        else:
            for cat in INCENTIVE_CATS:
                totals[f"GotMsg_{cat}"] = 0

        if commfree_col in df.columns:
            cf_all = pd.to_numeric(df[commfree_col], errors="coerce")
            totals["Free Comm Drivers"] = int((cf_all > 0).sum())
        else:
            totals["Free Comm Drivers"] = 0

        if n_total >= MIN_N and n_total > 0:
            totals["% Who Got Message"] = n_got_msg_total / n_total * 100
            if n_got_msg_total >= MIN_N:
                fc_msg_t = totals.get("GotMsg_Free-Commission", 0) + totals.get("GotMsg_Money & Free-commission", 0)
                totals["% Free Comm Message"] = fc_msg_t / n_got_msg_total * 100
            if filter_joint:
                if commfree_col in all_base.columns:
                    cf_base = pd.to_numeric(all_base[commfree_col], errors="coerce")
                    totals["% Free Comm Ride"] = int((cf_base > 0).sum()) / n_total * 100
                else:
                    totals["% Free Comm Ride"] = 0
            else:
                totals["% Free Comm Ride"] = totals["Free Comm Drivers"] / n_total * 100

        if part_col in got_msg_total.columns and not got_msg_total.empty:
            part_total = got_msg_total[got_msg_total[part_col] == "Yes"]
            for cat in INCENTIVE_CATS:
                cat_got = got_msg_total[got_msg_total[cat_col] == cat] if cat_col in got_msg_total.columns else pd.DataFrame()
                n_c = len(cat_got)
                if n_c >= MIN_N:
                    n_cp = len(part_total[part_total[cat_col] == cat]) if cat_col in part_total.columns else 0
                    totals[f"Participation_{cat}"] = n_cp / n_c * 100
            if n_got_msg_total >= MIN_N:
                totals["Participation_All"] = len(part_total) / n_got_msg_total * 100

        if ride_col in df.columns:
            ride_all = pd.to_numeric(df[ride_col], errors="coerce")
            total_rides_all = ride_all.sum()
            totals["Total Rides"] = total_rides_all
            if commfree_col in df.columns:
                cf_all_v = pd.to_numeric(df[commfree_col], errors="coerce")
                has_fc_all = cf_all_v > 0
                rides_fc_all = ride_all[has_fc_all].sum()
                fc_rides_all = cf_all_v[has_fc_all].sum()
                totals["RidesAmong_Total Rides"] = rides_fc_all
                totals["RidesAmong_Free Comm Rides"] = fc_rides_all
                if total_rides_all > 0:
                    totals["FreeCommShare_Among All"] = fc_rides_all / total_rides_all * 100
                if rides_fc_all > 0:
                    totals["FreeCommShare_Among Free Comm"] = fc_rides_all / rides_fc_all * 100

        total_row = pd.DataFrame([totals], index=["Total"])
        total_row.index.name = "City"
        result = pd.concat([result, total_row])
        # Enforce column order matching reference Excel layout
        col_order = [
            "n", "Who Got Message",
            "GotMsg_Money", "GotMsg_Free-Commission",
            "GotMsg_Money & Free-commission",
            "Free Comm Drivers",
            "% Who Got Message", "% Free Comm Message", "% Free Comm Ride",
            "Participation_Money", "Participation_Free-Commission",
            "Participation_Money & Free-commission", "Participation_All",
            "Total Rides",
            "RidesAmong_Total Rides", "RidesAmong_Free Comm Rides",
            "FreeCommShare_Among All", "FreeCommShare_Among Free Comm",
        ]
        ordered = [c for c in col_order if c in result.columns]
        extras = [c for c in result.columns if c not in col_order]
        result = result[ordered + extras]
        result = sort_cities(result)
        return result

    sm_week = filter_week(data["short_main"], week)
    results = {}

    tapsi_df = _build_commfree(sm_week, "tapsi")
    if not tapsi_df.empty:
        results["Tapsi"] = tapsi_df

    snapp_df = _build_commfree(sm_week, "snapp")
    if not snapp_df.empty:
        results["Snapp"] = snapp_df

    return results


def analysis_lucky_wheel(data, week):
    print("\n[#19] Lucky Wheel (Tapsi)...")
    df = filter_top_cities(filter_week(data["short_main"], week))
    if df.empty:
        return pd.DataFrame()
    res = pd.DataFrame(index=sorted(df["city"].unique()))
    if "wheel" in df.columns:
        res["wheel_usage"] = df.groupby("city")["wheel"].mean()
    if "tapsi_magical_window" in df.columns:
        mw = df[df["tapsi_magical_window"].notna()]
        if not mw.empty:
            ct = pd.crosstab(
                mw["city"], mw["tapsi_magical_window"], normalize="index") * 100
            ct = ct.round(1)
            res = res.join(ct)
    if "tapsi_magical_window_income" in df.columns:
        df["tapsi_magical_window_income"] = pd.to_numeric(
            df["tapsi_magical_window_income"], errors="coerce")
        res["avg_magical_window_income"] = df.groupby(
            "city")["tapsi_magical_window_income"].mean().round(0)
    res["n"] = df.groupby("city").size()
    res.index.name = "City"
    return res.dropna(how="all", axis=1)


def analysis_request_refusal(data, week):
    print("\n[Extra] Request Refusal Reasons...")
    lr = filter_top_cities(add_city(filter_week(
        data["long_rare"], week), data["_lookup"]))
    results = {}
    for q in ["Snapp Request Refusal", "Tapsi Request Refusal"]:
        qdf = lr[lr["question"] == q]
        if qdf.empty:
            continue
        resp = qdf.groupby("city")["recordID"].nunique()
        ans = qdf.groupby(["city", "answer"]).size().unstack(fill_value=0)
        pct = (ans.div(resp, axis=0) * 100).round(1)
        pct["n"] = resp
        pct.index.name = "City"
        results[q] = pct
    return results


# ═══════════════════════════════════════════════════════════════════════════
#  NEW ANALYSIS FUNCTIONS — from short_rare / wide_rare / long_rare
# ═══════════════════════════════════════════════════════════════════════════

def analysis_cs_satisfaction(data, week):
    """Customer Support satisfaction scores by city (from short_rare)."""
    print("\n[#CS] Customer Support Satisfaction...")
    sr = add_city(filter_week(data["short_rare"], week), data["_lookup"])
    sr = filter_top_cities(sr)

    results = {}
    for platform, prefix in [("Snapp", "snapp_CS_"), ("Tapsi", "tapsi_CS_")]:
        sat_cols = [c for c in sr.columns
                    if c.startswith(prefix + "satisfaction_")
                    and c != f"{prefix}satisfaction_important_reason"]
        solved_col = f"{prefix.rstrip('_')}_solved" if platform == "Tapsi" else f"{prefix}solved"
        # Fix: tapsi columns use tapsi_CS_solved
        solved_col = f"{prefix}solved"

        if not sat_cols:
            continue

        # Mean satisfaction by city
        df_result = mean_by_city(sr, sat_cols, top_cities=False)

        # CS solved rate
        if solved_col in sr.columns:
            solved_rate = sr.groupby("city")[solved_col].apply(
                lambda x: (x.dropna() == "Yes").mean() *
                100 if len(x.dropna()) > 0 else np.nan
            ).round(1).rename("solved_%")
            df_result = df_result.join(solved_rate, how="left")

        # CS contacted rate (snapp_CS or tapsi_CS_)
        contact_col = "snapp_CS" if platform == "Snapp" else "tapsi_CS"
        if contact_col in sr.columns:
            contact_rate = sr.groupby("city")[contact_col].apply(
                lambda x: (x.dropna() == "Yes").mean() *
                100 if len(x.dropna()) > 0 else np.nan
            ).round(1).rename("contacted_%")
            df_result = df_result.join(contact_rate, how="left")

        results[platform] = df_result
    return results


def analysis_cs_categories(data, week):
    """Customer Support category distribution (from long_rare/wide_rare)."""
    print("\n[#CS_Cat] Customer Support Categories...")
    lr = filter_top_cities(add_city(filter_week(
        data["long_rare"], week), data["_lookup"]))
    results = {}
    for q in ["Snapp Customer Support Category", "Tapsi Customer Support Category"]:
        qdf = lr[lr["question"] == q]
        if qdf.empty:
            continue
        resp = qdf.groupby("city")["recordID"].nunique()
        ans = qdf.groupby(["city", "answer"]).size().unstack(fill_value=0)
        pct = (ans.div(resp, axis=0) * 100).round(1)
        pct["n_contacted"] = resp
        pct.index.name = "City"
        results[q] = pct
    return results


def analysis_cs_important_reason(data, week):
    """Most important CS satisfaction factor (from short_rare)."""
    print("\n[#CS_Reason] CS Most Important Reason...")
    sr = add_city(filter_week(data["short_rare"], week), data["_lookup"])
    sr = filter_top_cities(sr)
    results = {}
    for platform, col in [
        ("Snapp", "snapp_CS_satisfaction_important_reason"),
        ("Tapsi", "tapsi_CS_satisfaction_important_reason"),
    ]:
        ct = crosstab_by_city(sr, col, top_cities=False)
        if not ct.empty:
            results[platform] = ct
    return results


def analysis_recommend(data, week):
    """Recommendation / NPS scores by city (from short_rare)."""
    print("\n[#Reco] Recommendation Scores...")
    sr = add_city(filter_week(data["short_rare"], week), data["_lookup"])
    sr = filter_top_cities(sr)

    # snapp_recommend and snappdriver_tapsi_recommend are likely 0-10 NPS
    reco_cols = ["snapp_recommend", "tapsidriver_tapsi_recommend",
                 "snappdriver_tapsi_recommend"]
    available = [c for c in reco_cols if c in sr.columns]
    if not available:
        return pd.DataFrame()

    for c in available:
        sr[c] = pd.to_numeric(sr[c], errors="coerce")

    result = mean_by_city(sr, available, top_cities=False)
    return result


def analysis_refer_others(data, week):
    """Would you refer others? distribution by city (from short_rare)."""
    print("\n[#Refer] Refer Others...")
    sr = add_city(filter_week(data["short_rare"], week), data["_lookup"])
    sr = filter_top_cities(sr)
    results = {}
    for col, label in [
        ("snapp_refer_others", "Snapp Refer Others"),
        ("tapsi_refer_others", "Tapsi Refer Others"),
    ]:
        ct = crosstab_by_city(sr, col, top_cities=False)
        if not ct.empty:
            results[label] = ct
    return results


def analysis_navigation_recommendations(data, week):
    """Navigation app recommendation scores (from short_rare)."""
    print("\n[#NavReco] Navigation App Recommendations...")
    sr = add_city(filter_week(data["short_rare"], week), data["_lookup"])
    sr = filter_top_cities(sr)

    reco_cols = [
        "recommendation_googlemap", "recommendation_waze",
        "recommendation_neshan", "recommendation_balad",
        "recommendation_googlemap_last3month", "recommendation_waze_last3month",
        "recommendation_neshan_last3month", "recommendation_balad_last3month",
        "snapp_navigation_app_satisfaction",
        "tapsi_in_app_navigation_satisfaction",
    ]
    available = [c for c in reco_cols if c in sr.columns]
    if not available:
        return pd.DataFrame()
    for c in available:
        sr[c] = pd.to_numeric(sr[c], errors="coerce")
    return mean_by_city(sr, available, top_cities=False)


def analysis_registration(data, week):
    """Registration type & motivation distribution by city (from short_rare)."""
    print("\n[#Reg] Registration Info...")
    sr = add_city(filter_week(data["short_rare"], week), data["_lookup"])
    sr = filter_top_cities(sr)
    results = {}
    for col, label in [
        ("snapp_register_type", "Snapp Register Type"),
        ("tapsi_register_type", "Tapsi Register Type"),
        ("snapp_main_reg_reason", "Snapp Main Reg Reason"),
        ("tapsi_main_reg_reason", "Tapsi Main Reg Reason"),
        ("tapsi_invite_to_reg", "Tapsi Invited to Register"),
    ]:
        ct = crosstab_by_city(sr, col, top_cities=False)
        if not ct.empty:
            results[label] = ct
    return results


def analysis_better_income(data, week):
    """Better income platform preference distribution by city (from short_rare)."""
    print("\n[#Income] Better Income Platform...")
    sr = add_city(filter_week(data["short_rare"], week), data["_lookup"])
    sr = filter_top_cities(sr)
    results = {}
    for col, label in [
        ("snapp_better_income", "Snapp Better Income"),
        ("tapsi_better_income", "Tapsi Better Income"),
    ]:
        ct = crosstab_by_city(sr, col, top_cities=False)
        if not ct.empty:
            results[label] = ct
    return results


def analysis_decline_reasons(data, week):
    """Decline/cancel reasons distribution (from long_rare)."""
    print("\n[#Decline] Decline Reasons...")
    lr = filter_top_cities(add_city(filter_week(
        data["long_rare"], week), data["_lookup"]))
    qdf = lr[lr["question"] == "Decline Reason"]
    if qdf.empty:
        return pd.DataFrame()
    resp = qdf.groupby("city")["recordID"].nunique()
    ans = qdf.groupby(["city", "answer"]).size().unstack(fill_value=0)
    pct = (ans.div(resp, axis=0) * 100).round(1)
    pct["n"] = resp
    pct.index.name = "City"
    return pct


def analysis_snappcarfix_satisfaction(data, week):
    """Snappcarfix satisfaction scores by city (from short_rare)."""
    print("\n[#Carfix] Snappcarfix Satisfaction...")
    sr = add_city(filter_week(data["short_rare"], week), data["_lookup"])
    sr = filter_top_cities(sr)

    sat_cols = [c for c in sr.columns if c.startswith(
        "snappcarfix_satisfaction_")]
    extra = ["snappcarfix_recommend"]
    all_cols = sat_cols + [c for c in extra if c in sr.columns]
    if not all_cols:
        return {}

    results = {}

    # Satisfaction scores
    for c in all_cols:
        sr[c] = pd.to_numeric(sr[c], errors="coerce")
    sat_result = mean_by_city(sr, all_cols, top_cities=False)
    if not sat_result.empty:
        results["Satisfaction"] = sat_result

    # Familiarity & usage rates
    for col, label in [
        ("snappcarfix_familiar", "Familiar"),
        ("snappcarfix_use_ever", "Used Ever"),
        ("snappcarfix_use_lastmo", "Used Last Month"),
    ]:
        ct = crosstab_by_city(sr, col, top_cities=False)
        if not ct.empty:
            results[label] = ct

    return results


def analysis_tapsigarage_satisfaction(data, week):
    """Tapsigarage satisfaction scores by city (from short_rare)."""
    print("\n[#Garage] Tapsigarage Satisfaction...")
    sr = add_city(filter_week(data["short_rare"], week), data["_lookup"])
    sr = filter_top_cities(sr)

    sat_cols = [c for c in sr.columns if c.startswith(
        "tapsigarage_satisfaction_")]
    extra = ["tapsigarage_recommend"]
    all_cols = sat_cols + [c for c in extra if c in sr.columns]
    if not all_cols:
        return {}

    results = {}
    for c in all_cols:
        sr[c] = pd.to_numeric(sr[c], errors="coerce")
    sat_result = mean_by_city(sr, all_cols, top_cities=False)
    if not sat_result.empty:
        results["Satisfaction"] = sat_result

    for col, label in [
        ("tapsigarage_familiar", "Familiar"),
        ("tapsigarage_use_ever", "Used Ever"),
        ("tapsigarage_use_lastmo", "Used Last Month"),
    ]:
        ct = crosstab_by_city(sr, col, top_cities=False)
        if not ct.empty:
            results[label] = ct
    return results


def analysis_demand(data, week):
    """Demand perception metrics by city (from short_rare)."""
    print("\n[#Demand] Demand Perception...")
    sr = add_city(filter_week(data["short_rare"], week), data["_lookup"])
    sr = filter_top_cities(sr)

    cols = ["max_demand", "demand_process", "missed_demand_per_10"]
    available = [c for c in cols if c in sr.columns]
    if not available:
        return pd.DataFrame()
    for c in available:
        sr[c] = pd.to_numeric(sr[c], errors="coerce")
    return mean_by_city(sr, available, top_cities=False)


def analysis_speed_satisfaction(data, week):
    """Speed satisfaction by city (from short_rare)."""
    print("\n[#Speed] Speed Satisfaction...")
    sr = add_city(filter_week(data["short_rare"], week), data["_lookup"])
    sr = filter_top_cities(sr)
    cols = ["snapp_speed_satisfaction", "tapsi_speed_satisfaction"]
    available = [c for c in cols if c in sr.columns]
    if not available:
        return pd.DataFrame()
    for c in available:
        sr[c] = pd.to_numeric(sr[c], errors="coerce")
    return mean_by_city(sr, available, top_cities=False)


def analysis_gps(data, week):
    """GPS problem and fix-location metrics by city (from short_rare)."""
    print("\n[#GPS] GPS Problems...")
    sr = add_city(filter_week(data["short_rare"], week), data["_lookup"])
    sr = filter_top_cities(sr)

    results = {}
    # GPS problem rate
    for col, label in [
        ("gps_problem", "GPS Problem"),
        ("gps_interrupt_impact", "GPS Interrupt Impact"),
        ("fixlocation_familiar", "FixLocation Familiar"),
        ("fixlocation_use", "FixLocation Use"),
        ("snapp_gps_stage", "Snapp GPS Stage"),
        ("tapsi_gps_stage", "Tapsi GPS Stage"),
        ("tapsi_gps_better", "Tapsi GPS Better"),
    ]:
        ct = crosstab_by_city(sr, col, top_cities=False)
        if not ct.empty:
            results[label] = ct

    # Satisfaction score
    if "fixlocation_satisfaction" in sr.columns:
        sr["fixlocation_satisfaction"] = pd.to_numeric(
            sr["fixlocation_satisfaction"], errors="coerce")
        sat = mean_by_city(sr, ["fixlocation_satisfaction"], top_cities=False)
        if not sat.empty:
            results["FixLoc Satisfaction"] = sat

    return results


def analysis_unpaid_by_passenger(data, week):
    """Unpaid by passenger metrics by city (from short_rare)."""
    print("\n[#Unpaid] Unpaid by Passenger...")
    sr = add_city(filter_week(data["short_rare"], week), data["_lookup"])
    sr = filter_top_cities(sr)

    results = {}
    # Unpaid rate
    ct = crosstab_by_city(sr, "unpaid_by_passenger", top_cities=False)
    if not ct.empty:
        results["Unpaid Rate"] = ct

    # Follow-up satisfaction
    for prefix, label in [("snapp_", "Snapp"), ("tapsi_", "Tapsi")]:
        sat_cols = [f"{prefix}satisfaction_followup_overall",
                    f"{prefix}satisfaction_followup_time"]
        available = [c for c in sat_cols if c in sr.columns]
        if available:
            for c in available:
                sr[c] = pd.to_numeric(sr[c], errors="coerce")
            sat = mean_by_city(sr, available, top_cities=False)
            if not sat.empty:
                results[f"{label} Followup Sat"] = sat

        # Compensated rate
        comp_col = f"{prefix}compensate_unpaid_by_passenger"
        ct2 = crosstab_by_city(sr, comp_col, top_cities=False)
        if not ct2.empty:
            results[f"{label} Compensated"] = ct2

    return results


def analysis_distance_to_origin(data, week):
    """Distance-to-origin time satisfaction (from short_rare)."""
    print("\n[#DistOrigin] Distance to Origin Satisfaction...")
    sr = add_city(filter_week(data["short_rare"], week), data["_lookup"])
    sr = filter_top_cities(sr)
    cols = ["snapp_distancetooring_time_satisfaction",
            "tapsi_distancetooring_time_satisfaction"]
    available = [c for c in cols if c in sr.columns]
    if not available:
        return pd.DataFrame()
    for c in available:
        sr[c] = pd.to_numeric(sr[c], errors="coerce")
    return mean_by_city(sr, available, top_cities=False)


# ─── Main ────────────────────────────────────────────────────────────────────
def run_all(week=None):
    data = load_data()
    if week is None:
        week = get_latest_week(data["short_main"])
    print(f"\nRunning analyses for Week {week}")
    print("=" * 60)

    sheets = {}

    def _safe_single(sheet_name, func, *args, **kwargs):
        """Run an analysis function, skip on missing-column errors."""
        try:
            return func(*args, **kwargs)
        except (KeyError, TypeError, ValueError) as e:
            print(f"  [WARN] Skipping '{sheet_name}': {e}")
            return pd.DataFrame()

    def _safe_multi(prefix, func, *args, **kwargs):
        """Run an analysis function that returns a dict, skip on errors."""
        try:
            return func(*args, **kwargs)
        except (KeyError, TypeError, ValueError) as e:
            print(f"  [WARN] Skipping '{prefix}': {e}")
            return {}

    # ── Original analyses (from short_main / wide_main / long_main) ──
    sheets["#1_Snapp_Incentive_Amt"] = _safe_single(
        "#1", analysis_incentive_amounts_snapp, data, week)
    sheets["#2_Tapsi_Incentive_Amt"] = _safe_single(
        "#2", analysis_incentive_amounts_tapsi, data, week)

    for seg, df in _safe_multi("#3_Satisfaction", analysis_satisfaction_review, data, week).items():
        sheets[f"#3_Sat_{seg[:20]}"] = df

    sheets["#4_Incentive_Duration"] = _safe_single(
        "#4", analysis_incentive_time_limitation, data, week)

    sheets["#5_6_IncType"] = _safe_single(
        "#5_6", analysis_received_incentive_types, data, week)

    try:
        dissat = analysis_incentive_dissatisfaction(data, week)
        sheets["#8_Dissat"] = dissat["combined"]
        summary = dissat.get("summary")
        if isinstance(summary, pd.DataFrame) and not summary.empty:
            sheets["#9_Dissat_Sum"] = summary
        last_wk = dissat.get("summary_last_week")
        if isinstance(last_wk, pd.DataFrame) and not last_wk.empty:
            sheets["#9_Dissat_Sum_LastWk"] = last_wk
    except (KeyError, TypeError, ValueError) as e:
        print(f"  [WARN] Skipping '#8/#9 Incentive Dissatisfaction': {e}")

    sheets["#12_Cities_Overview"] = _safe_single(
        "#12", analysis_all_cities_overview, data, week)

    sheets["#13_RideShare"] = _safe_single(
        "#13", analysis_ride_share, data, week)

    for label, df in _safe_multi("#14_Nav", analysis_navigation_usage, data, week).items():
        sheets[f"#14_Nav_{label[:20]}"] = df

    for label, df in _safe_multi("#15_Persona", analysis_driver_persona, data, week).items():
        sheets[f"#15_Persona_{label[:16]}"] = df

    sheets["#15_Persona_PartTime"] = _safe_single(
        "#15_PartTime", analysis_driver_persona_parttime_rides, data, week)

    for label, df in _safe_multi("#16_Ref", analysis_referral_plan, data, week).items():
        sheets[f"#16_Ref_{label[:20]}"] = df

    sheets["#17_Inactivity"] = _safe_single(
        "#17", analysis_inactivity_before_incentive, data, week)
    for label, df in _safe_multi("#18_CommFree", analysis_commission_free, data, week).items():
        sheets[f"#18_CommFree_{label}"] = df
    sheets["#19_LuckyWheel"] = _safe_single(
        "#19", analysis_lucky_wheel, data, week)

    for label, df in _safe_multi("#20_Refusal", analysis_request_refusal, data, week).items():
        safe = label.replace(" ", "_")[:20]
        sheets[f"#20_Refusal_{safe}"] = df

    # ── NEW analyses (from short_rare / wide_rare / long_rare) ──
    for plat, df in _safe_multi("#CS_Sat", analysis_cs_satisfaction, data, week).items():
        sheets[f"#CS_Sat_{plat}"] = df

    for label, df in _safe_multi("#CS_Cat", analysis_cs_categories, data, week).items():
        safe = label.replace(" ", "_")[:18]
        sheets[f"#CS_Cat_{safe}"] = df

    for plat, df in _safe_multi("#CS_Reason", analysis_cs_important_reason, data, week).items():
        sheets[f"#CS_Reason_{plat}"] = df

    reco = _safe_single("#Reco_NPS", analysis_recommend, data, week)
    if isinstance(reco, pd.DataFrame) and not reco.empty:
        sheets["#Reco_NPS"] = reco

    for label, df in _safe_multi("#Refer", analysis_refer_others, data, week).items():
        safe = label.replace(" ", "_")[:18]
        sheets[f"#Refer_{safe}"] = df

    nav_reco = _safe_single(
        "#NavReco", analysis_navigation_recommendations, data, week)
    if isinstance(nav_reco, pd.DataFrame) and not nav_reco.empty:
        sheets["#NavReco_Scores"] = nav_reco

    for label, df in _safe_multi("#Reg", analysis_registration, data, week).items():
        safe = label.replace(" ", "_")[:18]
        sheets[f"#Reg_{safe}"] = df

    for label, df in _safe_multi("#Income", analysis_better_income, data, week).items():
        safe = label.replace(" ", "_")[:18]
        sheets[f"#Income_{safe}"] = df

    decline = _safe_single("#Decline", analysis_decline_reasons, data, week)
    if isinstance(decline, pd.DataFrame) and not decline.empty:
        sheets["#Decline_Reasons"] = decline

    for label, df in _safe_multi("#Carfix", analysis_snappcarfix_satisfaction, data, week).items():
        sheets[f"#Carfix_{label[:20]}"] = df

    for label, df in _safe_multi("#Garage", analysis_tapsigarage_satisfaction, data, week).items():
        sheets[f"#Garage_{label[:20]}"] = df

    demand = _safe_single("#Demand", analysis_demand, data, week)
    if isinstance(demand, pd.DataFrame) and not demand.empty:
        sheets["#Demand_Perception"] = demand

    speed = _safe_single("#Speed", analysis_speed_satisfaction, data, week)
    if isinstance(speed, pd.DataFrame) and not speed.empty:
        sheets["#Speed_Satisfaction"] = speed

    dist = _safe_single("#DistOrigin", analysis_distance_to_origin, data, week)
    if isinstance(dist, pd.DataFrame) and not dist.empty:
        sheets["#DistOrigin_Sat"] = dist

    for label, df in _safe_multi("#GPS", analysis_gps, data, week).items():
        safe = label.replace(" ", "_")[:18]
        sheets[f"#GPS_{safe}"] = df

    for label, df in _safe_multi("#Unpaid", analysis_unpaid_by_passenger, data, week).items():
        safe = label.replace(" ", "_")[:18]
        sheets[f"#Unpaid_{safe}"] = df

    # ─── Clean up: drop all-NaN columns & fully-empty sheets ────────────
    cleaned = {}
    for name, df in sheets.items():
        if df is None or (isinstance(df, pd.DataFrame) and df.empty):
            continue
        # Drop columns where every value is NaN (rotating questions not asked this week)
        meta_cols = {"n", "n_joint", "n_dissatisfied", "n_contacted"}
        drop_cols = {c for c in df.columns
                     if c not in meta_cols and df[c].isna().all()}
        keep_cols = [c for c in df.columns if c not in drop_cols]
        df = df[keep_cols]

        # If only meta columns remain (all data cols were NaN), skip sheet
        remaining_data = [c for c in df.columns if c not in meta_cols]
        if not remaining_data:
            print(
                f"  Skipping '{name}': no data for this week (rotating question)")
            continue
        cleaned[name] = df

    # ─── Export with formatting ──────────────────────────────────────────
    output_path = os.path.join(
        OUTPUT_DIR, f"routine_analysis_week_{week}.xlsx")
    # If file is locked (open in Excel), try an alternate name
    if os.path.exists(output_path):
        try:
            with open(output_path, "a"):
                pass
        except PermissionError:
            alt = os.path.join(
                OUTPUT_DIR, f"routine_analysis_week_{week}_new.xlsx")
            print(f"\n  File is open in Excel, saving to: {alt}")
            output_path = alt
    print(f"\n{'=' * 60}")
    print(f"Exporting to {output_path}...")

    # Border & header styles
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(bold=True, size=9)
    header_fill = PatternFill(start_color="D9E1F2",
                              end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", wrap_text=True)
    data_font = Font(size=9)
    data_align = Alignment(horizontal="center")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, df in cleaned.items():
            safe_name = name[:31]

            df = sort_cities(df)
            df = reorder_columns(df, name)
            df = apply_min_n_cutoff(df, name)
            df = convert_pct_to_decimal(df.copy(), name)

            try:
                df.to_excel(writer, sheet_name=safe_name)
            except Exception as e:
                print(f"  Warning: sheet '{safe_name}': {e}")
                continue

            ws = writer.sheets[safe_name]

            # 0) Insert merged group-header row if defined for this sheet
            grp_headers = None
            for prefix, headers in SHEET_GROUP_HEADERS.items():
                if name.startswith(prefix):
                    grp_headers = headers
                    break
            has_group_header = grp_headers is not None
            if has_group_header:
                add_group_header_row(ws, df, thin_border, grp_headers)
                # Strip group prefixes from column header cells (row 2)
                # so headers read "Few Hours" instead of "Snapp_Few Hours"
                for prefix, _label in grp_headers:
                    for col_idx in range(2, ws.max_column + 1):
                        cell = ws.cell(row=2, column=col_idx)
                        if isinstance(cell.value, str) and cell.value.startswith(prefix):
                            cell.value = cell.value[len(prefix):]
                # Rename count columns to display names
                _col_display = {"E_n": "# Respondents", "F_n": "# Joint",
                                "G_n": "# Cmpt Signup"}
                hdr_r = 2  # header row when group header is present
                for col_idx in range(2, ws.max_column + 1):
                    cell = ws.cell(row=hdr_r, column=col_idx)
                    if cell.value in _col_display:
                        cell.value = _col_display[cell.value]

            # Strip "_WoW" → "WoW" and "_Last" → "Last" in column headers
            _suffix_row = 2 if has_group_header else 1
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=_suffix_row, column=col_idx)
                if isinstance(cell.value, str):
                    if cell.value.endswith("_WoW"):
                        cell.value = "WoW"
                    elif cell.value.endswith("_Last"):
                        cell.value = "Last"

            # Row offsets: with group header, col headers are row 2,
            # data starts row 3.  Without: headers row 1, data row 2.
            hdr_row = 2 if has_group_header else 1
            data_start = hdr_row + 1

            # 1) Borders on all cells + header formatting
            max_row = ws.max_row
            max_col = ws.max_column
            for row in ws.iter_rows(min_row=1, max_row=max_row,
                                    min_col=1, max_col=max_col):
                for cell in row:
                    cell.border = thin_border
                    if cell.row == hdr_row:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                    elif cell.row >= data_start:
                        cell.font = data_font
                        if cell.column > 1:  # skip city column
                            cell.alignment = data_align

            # 2) Conditional formatting (color scales)
            apply_conditional_formatting(ws, df, name,
                                         data_start_row=data_start)

            # 3) Percentage number format
            format_pct_cells(ws, df, name, data_start_row=data_start)

            # 4) Auto-fit column widths (compact heatmap style)
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                max_data_len = 0
                # Measure only data cells (skip header/group-header rows)
                for row_idx in range(data_start, min(max_row + 1, 50)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    if val is None:
                        continue
                    if cell.number_format == '0.0%' and isinstance(val, (int, float)):
                        rendered = f"{val * 100:.1f}%"
                    elif isinstance(val, float):
                        rendered = f"{val:.1f}" if val == int(
                            val) else f"{val:g}"
                    else:
                        rendered = str(val)
                    max_data_len = max(max_data_len, len(rendered))
                # Column A (city names): fit to content; others: tight
                if col_idx == 1:
                    # Measure header too for city column
                    for row_idx in range(1, data_start):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value:
                            max_data_len = max(max_data_len,
                                               len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_data_len + 1
                else:
                    ws.column_dimensions[col_letter].width = max(
                        max_data_len + 0.3, 2.5)

            # 4b) Compact row heights – taller headers for wrapped text
            for row_idx in range(1, max_row + 1):
                if row_idx < data_start:
                    # Header / group-header rows: allow 2-line wrap
                    ws.row_dimensions[row_idx].height = 30
                else:
                    ws.row_dimensions[row_idx].height = 15

            # 5) Freeze header rows + index column
            ws.freeze_panes = f"B{data_start}"

    print(f"Done! {len(cleaned)} sheets written to: {output_path}")
    return cleaned


def resolve_week(args, df):
    """
    Parse CLI args to a weeknumber.
    Accepts:  '2025 52'  or  '52'  or  nothing (auto-detect latest).
    When year+week given, maps to the weeknumber in the data.
    """
    if not args:
        return get_latest_week(df)

    if len(args) == 2:
        year, iso_week = int(args[0]), int(args[1])
    elif len(args) == 1:
        val = int(args[0])
        if val > 100:
            # e.g. "202552" → year=2025, week=52
            year, iso_week = divmod(val, 100)
        else:
            return val  # plain weeknumber
    else:
        print("Usage: python survey_routine_analysis.py [YEAR WEEK]")
        print("  e.g.: python survey_routine_analysis.py 2025 52")
        print("        python survey_routine_analysis.py 52")
        raise SystemExit(1)

    # Map year+ISO-week to the weeknumber in the data
    dt = df["datetime"].dropna()
    dt = pd.to_datetime(dt, errors="coerce")
    df_copy = pd.DataFrame({"weeknumber": df["weeknumber"], "dt": dt}).dropna()
    df_copy["iso_year"] = df_copy["dt"].dt.isocalendar().year.astype(int)
    df_copy["iso_week"] = df_copy["dt"].dt.isocalendar().week.astype(int)

    match = df_copy[(df_copy["iso_year"] == year) &
                    (df_copy["iso_week"] == iso_week)]
    if match.empty:
        print(f"No data found for year {year}, week {iso_week}")
        # Show available year-week combos
        available = df_copy.groupby(["iso_year", "iso_week"])[
            "weeknumber"].first()
        print(f"Available: {available.tail(10).to_dict()}")
        raise SystemExit(1)

    resolved = match["weeknumber"].mode().iloc[0]
    print(f"Resolved {year}-W{iso_week:02d} -> weeknumber {resolved}")
    return resolved


if __name__ == "__main__":
    import sys
    # Quick-load just the columns needed for week resolution
    _sm = pd.read_csv(os.path.join(PROCESSED_DIR, "short_survey_main.csv"),
                      usecols=["weeknumber", "datetime"], low_memory=False)
    week = resolve_week(sys.argv[1:], _sm)
    del _sm
    run_all(week)
