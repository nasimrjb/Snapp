"""
Header Frequency Analysis
Reads all .xlsx files in the raw survey folder, counts how many times
each column header appears across files, then compares against the
known questions in column_rename.xlsx and produces a formatted report.
"""

import os
import re
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIGURATION ────────────────────────────────────────────────────────────
RAW_DIR = r"D:\Work\Driver Survey\raw"
RENAME_FILE = r"D:\Work\Driver Survey\DataSources\column_rename.xlsx"
OUTPUT_FILE = r"D:\Work\Driver Survey\DataSources\header_frequency_report.xlsx"

# ── Persian normalisation (same logic as your DS_cleaning pipeline) ──────────
_AR_TO_FA = str.maketrans({
    "\u0643": "\u06a9", "\u064a": "\u06cc", "\u0629": "\u0647",
    "\u0624": "\u0648", "\u0625": "\u0627", "\u0623": "\u0627",
    "\u0671": "\u0627", "\u200c": " ", "\u200b": "", "\u200f": "",
    "\u200e": "", "\ufeff": "",
})
_DIACRITICS = re.compile(
    r"[\u064b-\u065f\u0610-\u061a\u06d6-\u06dc\u06df-\u06e4\u06e7\u06e8\u06ea-\u06ed]"
)


def fa_norm(text: str) -> str:
    text = str(text).translate(_AR_TO_FA)
    text = _DIACRITICS.sub("", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


# ── STEP 1: Read all headers from raw .xlsx files ───────────────────────────
raw_path = Path(RAW_DIR)
xlsx_files = sorted(raw_path.glob("*.xlsx"))
print(f"Found {len(xlsx_files)} .xlsx files in {RAW_DIR}")

header_counter = Counter()          # normalised header -> count
header_raw_examples = {}            # normalised header -> first raw text seen
files_with_header = {}              # normalised header -> set of filenames

for fpath in xlsx_files:
    fname = fpath.name
    if fname.startswith("~$"):
        continue
    try:
        xls = pd.ExcelFile(fpath, engine="openpyxl")
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet, nrows=0)
            for col in df.columns:
                raw = str(col).strip()
                norm = fa_norm(raw)
                if not norm:
                    continue
                header_counter[norm] += 1
                if norm not in header_raw_examples:
                    header_raw_examples[norm] = raw
                files_with_header.setdefault(norm, set()).add(fname)
    except Exception as e:
        print(f"  ⚠ Skipped {fname}: {e}")

print(f"Found {len(header_counter)} unique headers (after normalisation)")

# ── STEP 2: Load known questions from column_rename.xlsx ─────────────────────
qs = pd.read_excel(RENAME_FILE, sheet_name="questions")
# Column A = question_raw, Column B = question_short
known_raw_col = qs.columns[0]
known_short_col = qs.columns[1]

known_questions_raw = qs[known_raw_col].dropna().astype(
    str).str.strip().tolist()
known_questions_short = qs[known_short_col].tolist()
known_questions_norm = [fa_norm(q) for q in known_questions_raw]

# Build ordered maps: norm -> raw text, norm -> question_short
known_map = {}        # norm -> raw
known_short_map = {}  # norm -> question_short
for raw, short, norm in zip(known_questions_raw, known_questions_short, known_questions_norm):
    if norm not in known_map:
        known_map[norm] = raw
        known_short_map[norm] = str(short).strip() if pd.notna(short) else ""

print(f"Loaded {len(known_map)} known questions from column_rename.xlsx")

# ── STEP 3: Build the output rows ───────────────────────────────────────────
rows = []
seen_norms = set()

# First: all questions from column_rename in their original order
for norm in known_questions_norm:
    if norm in seen_norms:
        continue
    seen_norms.add(norm)
    freq = header_counter.get(norm, 0)
    raw_text = known_map[norm]
    short_name = known_short_map.get(norm, "")
    rows.append((raw_text, short_name, freq, "Available"))

# Second: headers found in raw files but NOT in column_rename
for norm, count in header_counter.most_common():
    if norm not in seen_norms:
        seen_norms.add(norm)
        raw_text = header_raw_examples[norm]
        rows.append((raw_text, "", count, "Not filed"))

# ── STEP 4: Create formatted output xlsx ────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Header Frequency"

# Styles
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F5496")
header_align = Alignment(
    horizontal="center", vertical="center", wrap_text=True)

avail_fill = PatternFill("solid", fgColor="E2EFDA")
not_filed_fill = PatternFill("solid", fgColor="FCE4EC")
zero_font = Font(name="Arial", color="C00000", size=10)
normal_font = Font(name="Arial", size=10)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Headers
headers = ["Question", "Question Short", "Frequency", "Status"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Data rows
for r, (question, short_name, freq, status) in enumerate(rows, 2):
    c1 = ws.cell(row=r, column=1, value=question)
    c2 = ws.cell(row=r, column=2, value=short_name)
    c3 = ws.cell(row=r, column=3, value=freq)
    c4 = ws.cell(row=r, column=4, value=status)

    c1.font = normal_font
    c1.alignment = Alignment(vertical="center", wrap_text=True)
    c2.font = normal_font
    c2.alignment = Alignment(vertical="center")
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c4.alignment = Alignment(horizontal="center", vertical="center")

    if freq == 0:
        c3.font = zero_font
    else:
        c3.font = normal_font

    if status == "Available":
        c4.fill = avail_fill
        c4.font = Font(name="Arial", size=10, color="375623")
    else:
        c4.fill = not_filed_fill
        c4.font = Font(name="Arial", size=10, color="C00000")

    for cell in (c1, c2, c3, c4):
        cell.border = thin_border

# Column widths
ws.column_dimensions["A"].width = 65
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 14

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:D{len(rows) + 1}"

# ── Summary sheet ────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
summary_data = [
    ("Total .xlsx files scanned", len(xlsx_files)),
    ("Unique headers found", len(header_counter)),
    ("Known questions (column_rename)", len(known_map)),
    ("Available (matched)", sum(1 for _, _, _, s in rows if s == "Available")),
    ("Not filed (new headers)", sum(1 for _, _, _, s in rows if s == "Not filed")),
    ("Available with 0 frequency", sum(1 for _, _, f,
     s in rows if s == "Available" and f == 0)),
]
for r, (label, val) in enumerate(summary_data, 1):
    ws2.cell(row=r, column=1, value=label).font = Font(
        name="Arial", bold=True, size=10)
    ws2.cell(row=r, column=2, value=val).font = Font(name="Arial", size=10)
    ws2.cell(row=r, column=2).alignment = Alignment(horizontal="center")
ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 14

wb.save(OUTPUT_FILE)
print(f"\n✅ Report saved to: {OUTPUT_FILE}")
print(f"   {sum(1 for _, _, _, s in rows if s == 'Available')} Available | "
      f"{sum(1 for _, _, _, s in rows if s == 'Not filed')} Not filed | "
      f"{sum(1 for _, _, f, s in rows if s == 'Available' and f == 0)} Available with 0 frequency")
