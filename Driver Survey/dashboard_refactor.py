"""
╔══════════════════════════════════════════════════════════════════════════════╗
║          TAPSI & SNAPP DRIVER SURVEY — ANALYSIS SCRIPT                      ║
║          Generates a Power BI-ready Excel workbook with heatmap matrices     ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  HOW TO USE                                                                  ║
║  ─────────────────────────────────────────────────────────────────────────  ║
║  1. Install dependencies (once):                                             ║
║       pip install pandas openpyxl                                            ║
║                                                                              ║
║  2. Set your file paths in the CONFIG section below (around line 60)         ║
║                                                                              ║
║  3. Run:                                                                     ║
║       python survey_analysis.py                                              ║
║                                                                              ║
║  INPUT FILES REQUIRED                                                        ║
║  ─────────────────────────────────────────────────────────────────────────  ║
║  A) survey_short_format_single.csv  — one row per driver, all single-answer  ║
║     questions. Must contain at minimum these columns:                        ║
║       recordID, age, cooperation_type, active_time, ride_snapp, ride_tapsi, ║
║       overall_satisfaction_snapp, fare_satisfaction_snapp,                   ║
║       req_count_satisfaction_snapp, income_satisfaction_snapp,               ║
║       overall_satisfaction_tapsi, fare_satisfaction_tapsi,                   ║
║       req_count_satisfaction_tapsi, income_satisfaction_tapsi,               ║
║       incentive_category_snapp, incentive_category_tapsi,                   ║
║       overall_incentive_satisfaction_snapp,                                  ║
║       overall_incentive_satisfaction_tapsi,                                  ║
║       incentive_got_message_snapp, recommend_snapp, recommend_tapsi,         ║
║       demand_process, missed_demand_per_10                                   ║
║                                                                              ║
║  B) survey_long_format_multichoice.csv — one row per answer to a            ║
║     multi-choice question. Must contain:                                     ║
║       recordID, main_question, sub_question, answer                          ║
║     main_question values used:                                               ║
║       'decline_reason', 'incentive_type_snapp', 'incentive_type_tapsi'      ║
║                                                                              ║
║  OUTPUT                                                                      ║
║  ─────────────────────────────────────────────────────────────────────────  ║
║  Excel file with 9 color-coded sheets:                                       ║
║    📋 How to Use        — guide & legend                                     ║
║    Data_Drivers         — clean fact table (import into Power BI)            ║
║    Data_MultiChoice     — long-format table (import into Power BI)           ║
║    Matrix_Satisfaction  — heatmap: satisfaction by age × driver type         ║
║    Matrix_Incentive     — heatmap: incentive breakdown by segment            ║
║    Matrix_Behavior      — heatmap: NPS, demand, decline reasons              ║
║    Matrix_Tapsi         — heatmap: Tapsi-only (dual-platform drivers)        ║
║    Summary_Stats        — KPI summary                                        ║
║    Matrix_CrossPlatform — Snapp vs Tapsi side-by-side                        ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import sys
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG — CHANGE THESE PATHS TO MATCH YOUR FILES
# ══════════════════════════════════════════════════════════════════════════════

SHORT_CSV  = r"D:\OneDrive\Work\Driver Survey\Outputs\survey_short_format_single.csv"       # ← path to your short-format file
# ← path to your long-format file
LONG_CSV = r"D:\OneDrive\Work\Driver Survey\Outputs\survey_long_format_multichoice.csv"
# ← where the output will be saved
OUTPUT_XLSX = r"D:\OneDrive\Work\Driver Survey\Outputs\Driver_Survey_Dashboard.xlsx"

# ══════════════════════════════════════════════════════════════════════════════
# LOAD & VALIDATE DATA
# ══════════════════════════════════════════════════════════════════════════════

print("Loading data...")
try:
    short = pd.read_csv(SHORT_CSV)
    long  = pd.read_csv(LONG_CSV)
except FileNotFoundError as e:
    print(f"\n❌ ERROR: Could not find input file → {e}")
    print("   Please update SHORT_CSV and LONG_CSV paths at the top of this script.")
    sys.exit(1)

print(f"  ✓ Short format: {len(short):,} rows, {len(short.columns)} columns")
print(f"  ✓ Long format:  {len(long):,} rows, {len(long.columns)} columns")

# ══════════════════════════════════════════════════════════════════════════════
# DATA ENRICHMENT — derive helper columns
# ══════════════════════════════════════════════════════════════════════════════

# Flag whether driver has used Tapsi (ride_tapsi is non-null = they used Tapsi)
short['has_tapsi'] = short['ride_tapsi'].notna().map({True: 'Used Tapsi', False: 'Snapp Only'})

# Human-readable age labels
short['age_label'] = short['age'].map({
    '<18':   '<18',
    '18_25': '18-25',
    '26_35': '26-35',
    '36_45': '36-45',
    '46_55': '46-55',
    '56_65': '56-65',
    '>65':   '>65',
})

# Ordered age category list (controls row order in matrix tables)
AGE_ORDER = ['18-25', '26-35', '36-45', '46-55', '56-65', '>65']

# Map raw active_time strings to grouped activity levels
active_map = {
    'few hours/month':  'Occasional (<1h/day)',
    '<20hour/mo':       'Occasional (<1h/day)',
    '5_20hour/week':    'Part-Time (5-20h/wk)',
    '20_40h/week':      'Part-Time (5-20h/wk)',
    '8_12hour/day':     'Full-Time (8-12h/day)',
    '>40h/week':        'Full-Time (>40h/wk)',
    '>12h/day':         'Full-Time (>12h/day)',
}
short['activity_group'] = short['active_time'].map(active_map)

# Ordered driver type (controls column order in matrix tables)
DRIVER_ORDER = ['Part-Time', 'Full-Time']
short['cooperation_type'] = pd.Categorical(
    short['cooperation_type'], categories=DRIVER_ORDER, ordered=True
)

# Incentive category display order
INC_ORDER = ['Money', 'Free-Commission', 'Money & Free-commission']

# Subset of drivers who used BOTH Snapp and Tapsi
tapsi = short[short['has_tapsi'] == 'Used Tapsi'].copy()

print(f"  ✓ Snapp-only drivers: {(short['has_tapsi']=='Snapp Only').sum():,}")
print(f"  ✓ Dual-platform drivers (used Tapsi too): {len(tapsi):,}")

# ══════════════════════════════════════════════════════════════════════════════
# STYLE CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

NAVY     = 'FF1F3864'   # dark navy — main headers
BLUE     = 'FF2E75B6'   # medium blue — sub-headers
SNAPP    = 'FF1A5276'   # dark teal — Snapp sections
TAPSI    = 'FF922B21'   # dark red — Tapsi sections
BOTH     = 'FF1E8449'   # green — cross-platform / instructions
LGRAY    = 'FFF2F3F4'   # light gray — alternating row background
WHITE    = 'FFFFFFFF'
DARKTEXT = 'FF1C2833'

# ══════════════════════════════════════════════════════════════════════════════
# STYLE HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def h1(cell, text, bg=NAVY, fg=WHITE, size=13):
    """Large header cell (section title bars)."""
    cell.value = text
    cell.font = Font(bold=True, color=fg, size=size, name='Calibri')
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def h2(cell, text, bg=BLUE, fg=WHITE, size=11):
    """Medium header cell (column / row header)."""
    cell.value = text
    cell.font = Font(bold=True, color=fg, size=size, name='Calibri')
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def row_header(cell, text, bg=LGRAY):
    """Left-column row label in a matrix table."""
    cell.value = text
    cell.font = Font(bold=True, size=10, name='Calibri', color=DARKTEXT)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='left', vertical='center')


def data_cell(cell, val, bg=WHITE):
    """Numeric data cell (score / percentage value)."""
    if pd.notna(val) and val != '-':
        cell.value = round(float(val), 2)
        cell.number_format = '0.00'
    else:
        cell.value = '—'
    cell.font = Font(size=10, name='Calibri', color=DARKTEXT)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')


def count_cell(cell, val, bg=WHITE):
    """Integer count data cell."""
    if pd.notna(val):
        cell.value = int(val)
        cell.number_format = '#,##0'
    else:
        cell.value = 0
    cell.font = Font(size=10, name='Calibri', color=DARKTEXT)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')


def thin_border():
    s = Side(style='thin', color='FFADB5BD')
    return Border(left=s, right=s, top=s, bottom=s)


def border_range(ws, r1, r2, c1, c2):
    """Apply thin border to a rectangular cell range."""
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for cell in row:
            cell.border = thin_border()


def heatmap(ws, r1, r2, c1, c2,
            lo='FFF8CBAD', mid='FFFFD966', hi='FF70AD47'):
    """Apply 3-color scale conditional formatting (red→yellow→green by default)."""
    rng = f'{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}'
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type='min',        start_color=lo,
        mid_type='percentile',   mid_value=50, mid_color=mid,
        end_type='max',          end_color=hi,
    ))


def section_title(ws, row, col, text, span, bg=NAVY, fg=WHITE, size=11):
    """Write a full-width section divider bar."""
    ws.cell(row, col).value = text
    ws.cell(row, col).font = Font(bold=True, color=fg, size=size, name='Calibri')
    ws.cell(row, col).fill = PatternFill('solid', start_color=bg)
    ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
    if span > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=col + span - 1
        )


def write_matrix(ws, start_row, start_col, pivot, title,
                 row_label='',
                 val_fmt='score',           # 'score' | 'count'
                 hm_lo='FFF8CBAD',          # heatmap low  color (hex, no #)
                 hm_mid='FFFFD966',         # heatmap mid  color
                 hm_hi='FF70AD47',          # heatmap high color
                 bg_header=BLUE):
    """
    Write a pandas pivot table as a formatted matrix with heatmap.

    Parameters
    ----------
    ws          : openpyxl worksheet
    start_row   : first row to write into (1-indexed)
    start_col   : first column to write into (1-indexed)
    pivot       : pd.DataFrame — index = row labels, columns = column labels
    title       : string shown in the title bar above the matrix
    row_label   : label for the top-left corner cell
    val_fmt     : 'score' writes floats; 'count' writes integers
    hm_lo/mid/hi: ARGB hex strings for the 3-color heatmap scale
    bg_header   : ARGB hex for column header background

    Returns
    -------
    int — the next available row after the matrix (with 1 blank row gap)
    """
    r, c = start_row, start_col
    cols = list(pivot.columns)
    total_cols = len(cols) + 1   # +1 for the row-label column

    # ── Title bar ─────────────────────────────────────────────────────────────
    ws.merge_cells(
        start_row=r, start_column=c,
        end_row=r,   end_column=c + total_cols - 1
    )
    h1(ws.cell(r, c), title, bg=NAVY, size=11)
    r += 1

    # ── Column headers ─────────────────────────────────────────────────────────
    h2(ws.cell(r, c), row_label or '↓ / →', bg=bg_header)
    for j, col in enumerate(cols):
        h2(ws.cell(r, c + 1 + j), str(col), bg=bg_header)
    r += 1

    data_start_row = r

    # ── Data rows ──────────────────────────────────────────────────────────────
    for i, idx in enumerate(pivot.index):
        bg = LGRAY if i % 2 == 0 else WHITE
        row_header(ws.cell(r, c), str(idx), bg=bg)
        for j, col in enumerate(cols):
            val = pivot.loc[idx, col]
            if val_fmt == 'count':
                count_cell(ws.cell(r, c + 1 + j), val, bg=bg)
            else:
                data_cell(ws.cell(r, c + 1 + j), val, bg=bg)
        r += 1

    data_end_row = r - 1

    # ── Heatmap conditional formatting (data area only, not headers) ───────────
    if data_end_row >= data_start_row:
        heatmap(ws, data_start_row, data_end_row,
                c + 1, c + len(cols),
                lo=hm_lo, mid=hm_mid, hi=hm_hi)

    # ── Borders & column widths ────────────────────────────────────────────────
    border_range(ws, start_row, data_end_row, c, c + total_cols - 1)
    ws.column_dimensions[get_column_letter(c)].width = 26
    for j in range(len(cols)):
        ws.column_dimensions[get_column_letter(c + 1 + j)].width = 16

    return data_end_row + 2   # +1 blank gap row


# ══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════

print("\nBuilding workbook...")
wb = Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 0 — How to Use (instructions & legend)
# ─────────────────────────────────────────────────────────────────────────────
ws0 = wb.active
ws0.title = '📋 How to Use'
ws0.sheet_view.showGridLines = False
ws0.column_dimensions['A'].width = 4
ws0.column_dimensions['B'].width = 34
ws0.column_dimensions['C'].width = 60

ws0.row_dimensions[1].height = 50
ws0.merge_cells('B1:C1')
h1(ws0['B1'],
   '🚖 Tapsi & Snapp Driver Survey — Power BI Analysis Pack',
   bg=NAVY, size=16)

info_rows = [
    ('', '', ''),
    ('', '📌 What is this file?',
     f'Power BI-ready analysis of {len(short):,} driver survey responses '
     f'comparing Snapp and Tapsi platforms.'),
    ('', '📂 Sheets in this workbook', ''),
    ('', '  ① Data_Drivers',      'Clean driver-level data → Import into Power BI as main fact table'),
    ('', '  ② Data_MultiChoice',  'Long-format multi-choice responses → Import as dimension table'),
    ('', '  ③ Matrix_Satisfaction','Heatmap matrices: satisfaction scores by age & driver type'),
    ('', '  ④ Matrix_Incentive',   'Heatmap matrices: incentive type breakdown by segment'),
    ('', '  ⑤ Matrix_Behavior',    'Demand acceptance, NPS, decline reasons'),
    ('', '  ⑥ Matrix_Tapsi',       'Tapsi-specific metrics (dual-platform drivers only)'),
    ('', '  ⑦ Summary_Stats',      'High-level KPIs and distributions'),
    ('', '  ⑧ Matrix_CrossPlatform','Snapp vs Tapsi side-by-side comparisons'),
    ('', '', ''),
    ('', '🎨 Heatmap Color Scale', ''),
    ('', '  🔴 Red / Orange',      'Low values (low satisfaction / low share)'),
    ('', '  🟡 Yellow',            'Mid-range values'),
    ('', '  🟢 Green',             'High values (high satisfaction / high share)'),
    ('', '', ''),
    ('', '⚡ Power BI Import Guide', ''),
    ('', '  Step 1', 'Open Power BI Desktop → Get Data → Excel Workbook'),
    ('', '  Step 2', 'Select this file → check ✅ Data_Drivers + Data_MultiChoice'),
    ('', '  Step 3', 'Load → build Matrix visual → enable Conditional Formatting (heatmap)'),
    ('', '  Step 4', 'Use the Matrix_* sheets as reference for layout & metrics'),
    ('', '', ''),
    ('', '📊 Key Segments', ''),
    ('', '  Driver Type',    f'Part-Time ({(short["cooperation_type"]=="Part-Time").sum():,}) '
                              f'vs Full-Time ({(short["cooperation_type"]=="Full-Time").sum():,})'),
    ('', '  Age Groups',     '18-25 / 26-35 / 36-45 / 46-55 / 56-65 / >65'),
    ('', '  Platform Use',   f'Snapp Only ({(short["has_tapsi"]=="Snapp Only").sum():,}) '
                              f'vs Both ({len(tapsi):,})'),
    ('', '  Incentive Type', 'Money | Free-Commission | Money & Free-Commission'),
]

for i, (_, label, val) in enumerate(info_rows):
    rr = i + 2
    ws0.row_dimensions[rr].height = 18
    cell_l = ws0.cell(rr, 2)
    cell_v = ws0.cell(rr, 3)
    cell_l.value = label
    cell_v.value = val
    is_section = any(label.startswith(p) for p in ['📌','📂','🎨','⚡','📊'])
    is_sheet   = any(label.startswith(f'  {n}') for n in ['①','②','③','④','⑤','⑥','⑦','⑧'])
    is_step    = label.startswith('  Step')
    if is_section:
        cell_l.font = Font(bold=True, size=11, color=NAVY, name='Calibri')
    elif is_sheet:
        cell_l.font = Font(bold=True, size=10, color='FF2E75B6', name='Calibri')
        cell_v.font = Font(size=10, name='Calibri')
    elif is_step:
        cell_l.font = Font(bold=True, size=10, color=BOTH, name='Calibri')
        cell_v.font = Font(size=10, name='Calibri')
    else:
        cell_l.font = Font(size=10, name='Calibri', color=DARKTEXT)
        cell_v.font = Font(size=10, name='Calibri', color=DARKTEXT)
    cell_l.alignment = Alignment(vertical='center')
    cell_v.alignment = Alignment(vertical='center', wrap_text=True)

print("  ✓ Sheet: How to Use")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 — Data_Drivers  (one row per driver — import this into Power BI)
# ─────────────────────────────────────────────────────────────────────────────
ws1 = wb.create_sheet('Data_Drivers')
ws1.freeze_panes = 'B2'

EXPORT_COLS = [
    'recordID', 'datetime', 'gender', 'age', 'age_label', 'age_group',
    'education', 'marital_status', 'cooperation_type', 'active_time',
    'activity_group', 'has_tapsi', 'ride_snapp', 'ride_tapsi',
    'overall_satisfaction_snapp', 'fare_satisfaction_snapp',
    'req_count_satisfaction_snapp', 'income_satisfaction_snapp',
    'overall_satisfaction_tapsi', 'fare_satisfaction_tapsi',
    'req_count_satisfaction_tapsi', 'income_satisfaction_tapsi',
    'incentive_category_snapp', 'incentive_category_tapsi',
    'overall_incentive_satisfaction_snapp', 'overall_incentive_satisfaction_tapsi',
    'incentive_got_message_snapp', 'incentive_got_message_tapsi',
    'incentive_rial_details_snapp', 'incentive_rial_details_tapsi',
    'recommend_snapp', 'recommend_tapsi',
    'demand_process', 'max_demand', 'missed_demand_per_10',
    'carpooling_familiar_tapsi', 'carpooling_gotoffer_accepted_tapsi',
    'carpooling_satisfaction_tapsi_overall',
    'register_type_snapp', 'main_reg_reason_snapp',
    'register_type_tapsi', 'main_reg_reason_tapsi',
    'familiar_accepted_percent_snapp',
    'diff_commfree_snapp', 'diff_commfree_tapsi',
    'commfree_snapp', 'commfree_tapsi',
    'age_snapp', 'age_tapsi', 'trip_count_snapp', 'trip_count_tapsi',
]

# Friendly column header labels shown in row 1
COL_LABELS = {
    'recordID': 'Record ID', 'datetime': 'Survey Date', 'gender': 'Gender',
    'age': 'Age Group (Raw)', 'age_label': 'Age Group',
    'age_group': 'Age Bracket (2-level)', 'education': 'Education',
    'marital_status': 'Marital Status', 'cooperation_type': 'Driver Type',
    'active_time': 'Active Hours (Raw)', 'activity_group': 'Activity Level',
    'has_tapsi': 'Platform Usage', 'ride_snapp': 'Snapp Rides',
    'ride_tapsi': 'Tapsi Rides',
    'overall_satisfaction_snapp': 'Overall Sat. Snapp (1-5)',
    'fare_satisfaction_snapp': 'Fare Sat. Snapp (1-5)',
    'req_count_satisfaction_snapp': 'Request Sat. Snapp (1-5)',
    'income_satisfaction_snapp': 'Income Sat. Snapp (1-5)',
    'overall_satisfaction_tapsi': 'Overall Sat. Tapsi (1-5)',
    'fare_satisfaction_tapsi': 'Fare Sat. Tapsi (1-5)',
    'req_count_satisfaction_tapsi': 'Request Sat. Tapsi (1-5)',
    'income_satisfaction_tapsi': 'Income Sat. Tapsi (1-5)',
    'incentive_category_snapp': 'Incentive Category Snapp',
    'incentive_category_tapsi': 'Incentive Category Tapsi',
    'overall_incentive_satisfaction_snapp': 'Incentive Sat. Snapp (1-5)',
    'overall_incentive_satisfaction_tapsi': 'Incentive Sat. Tapsi (1-5)',
    'incentive_got_message_snapp': 'Got Incentive Msg Snapp',
    'incentive_got_message_tapsi': 'Got Incentive Msg Tapsi',
    'incentive_rial_details_snapp': 'Incentive Amount Snapp (Rial)',
    'incentive_rial_details_tapsi': 'Incentive Amount Tapsi (Rial)',
    'recommend_snapp': 'NPS Snapp (0-10)', 'recommend_tapsi': 'NPS Tapsi (0-10)',
    'demand_process': 'Demand Acceptance Rate', 'max_demand': 'Max Concurrent Demand',
    'missed_demand_per_10': 'Missed Demands per 10',
    'carpooling_familiar_tapsi': 'Familiar with Tapsi Carpool',
    'carpooling_gotoffer_accepted_tapsi': 'Carpool Offer Accepted',
    'carpooling_satisfaction_tapsi_overall': 'Carpool Sat. Tapsi (1-5)',
    'register_type_snapp': 'Registration Type Snapp',
    'main_reg_reason_snapp': 'Reg. Reason Snapp',
    'register_type_tapsi': 'Registration Type Tapsi',
    'main_reg_reason_tapsi': 'Reg. Reason Tapsi',
    'familiar_accepted_percent_snapp': 'Demand Acceptance % Snapp',
    'diff_commfree_snapp': 'Commfree Diff Snapp',
    'diff_commfree_tapsi': 'Commfree Diff Tapsi',
    'commfree_snapp': 'Commfree Rides Snapp',
    'commfree_tapsi': 'Commfree Rides Tapsi',
    'age_snapp': 'Snapp Account Age', 'age_tapsi': 'Tapsi Account Age',
    'trip_count_snapp': 'Trip Count Snapp', 'trip_count_tapsi': 'Trip Count Tapsi',
}

df_export = short[[c for c in EXPORT_COLS if c in short.columns]].copy()

# Header row
for j, col in enumerate(df_export.columns):
    h2(ws1.cell(1, j + 1), COL_LABELS.get(col, col), bg=NAVY)
    ws1.column_dimensions[get_column_letter(j + 1)].width = 22
ws1.row_dimensions[1].height = 40

# Data rows (alternating row shading)
for i, (_, row) in enumerate(df_export.iterrows()):
    bg = LGRAY if i % 2 == 0 else WHITE
    for j, col in enumerate(df_export.columns):
        cell = ws1.cell(i + 2, j + 1)
        val = row[col]
        cell.value = val if pd.notna(val) else None
        cell.font = Font(size=9, name='Calibri')
        cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(vertical='center')

ws1.auto_filter.ref = f'A1:{get_column_letter(len(df_export.columns))}1'
print(f"  ✓ Sheet: Data_Drivers ({len(df_export):,} rows)")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 — Data_MultiChoice  (long format — import this into Power BI)
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Data_MultiChoice')
ws2.freeze_panes = 'A2'

long2 = long[['recordID', 'main_question', 'sub_question', 'answer']].merge(
    short[['recordID', 'age', 'age_label', 'cooperation_type', 'has_tapsi', 'activity_group']],
    on='recordID', how='left'
)

MC_LABELS = {
    'recordID': 'Record ID', 'main_question': 'Question Category',
    'sub_question': 'Sub Question', 'answer': 'Answer',
    'age': 'Age Raw', 'age_label': 'Age Group',
    'cooperation_type': 'Driver Type', 'has_tapsi': 'Platform Usage',
    'activity_group': 'Activity Level',
}

for j, col in enumerate(long2.columns):
    h2(ws2.cell(1, j + 1), MC_LABELS.get(col, col), bg=TAPSI)
    ws2.column_dimensions[get_column_letter(j + 1)].width = 30
ws2.row_dimensions[1].height = 35

for i, (_, row) in enumerate(long2.iterrows()):
    bg = LGRAY if i % 2 == 0 else WHITE
    for j, col in enumerate(long2.columns):
        cell = ws2.cell(i + 2, j + 1)
        val = row[col]
        cell.value = val if pd.notna(val) else None
        cell.font = Font(size=9, name='Calibri')
        cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(vertical='center')

ws2.auto_filter.ref = f'A1:{get_column_letter(len(long2.columns))}1'
print(f"  ✓ Sheet: Data_MultiChoice ({len(long2):,} rows)")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 — Matrix_Satisfaction
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet('Matrix_Satisfaction')
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 3

r = 2
ws3.merge_cells(f'B{r}:M{r}')
h1(ws3.cell(r, 2),
   '📊 SATISFACTION ANALYSIS — Snapp & Tapsi Drivers',
   bg=NAVY, size=14)
ws3.row_dimensions[r].height = 36
r += 2

# Shared heatmap colours for satisfaction (red=low, green=high)
SAT_LO, SAT_MID, SAT_HI = 'FFAE6161', 'FFFFD966', 'FF52BE80'

# ① Overall satisfaction by age × driver type
section_title(ws3, r, 2,
    '① SNAPP — Overall Satisfaction by Age Group & Driver Type (Avg 1-5)', 12, bg=SNAPP)
ws3.row_dimensions[r].height = 24
r += 1
piv = short.pivot_table(
    values='overall_satisfaction_snapp',
    index='age_label', columns='cooperation_type', aggfunc='mean'
).reindex([a for a in AGE_ORDER if a in short['age_label'].unique()])
r = write_matrix(ws3, r, 2, piv,
    'Snapp Overall Satisfaction (1-5)', row_label='Age Group',
    bg_header=SNAPP, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)
r += 1

# ② Fare satisfaction
section_title(ws3, r, 2,
    '② SNAPP — Fare Satisfaction by Age Group & Driver Type', 12, bg=SNAPP)
ws3.row_dimensions[r].height = 24
r += 1
piv = short.pivot_table(
    values='fare_satisfaction_snapp',
    index='age_label', columns='cooperation_type', aggfunc='mean'
).reindex([a for a in AGE_ORDER if a in short['age_label'].unique()])
r = write_matrix(ws3, r, 2, piv,
    'Snapp Fare Satisfaction (1-5)', row_label='Age Group',
    bg_header=SNAPP, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)
r += 1

# ③ Income satisfaction
section_title(ws3, r, 2,
    '③ SNAPP — Income Satisfaction by Age Group & Driver Type', 12, bg=SNAPP)
ws3.row_dimensions[r].height = 24
r += 1
piv = short.pivot_table(
    values='income_satisfaction_snapp',
    index='age_label', columns='cooperation_type', aggfunc='mean'
).reindex([a for a in AGE_ORDER if a in short['age_label'].unique()])
r = write_matrix(ws3, r, 2, piv,
    'Snapp Income Satisfaction (1-5)', row_label='Age Group',
    bg_header=SNAPP, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)
r += 1

# ④ Request count satisfaction
section_title(ws3, r, 2,
    '④ SNAPP — Request Count Satisfaction by Age Group & Driver Type', 12, bg=SNAPP)
ws3.row_dimensions[r].height = 24
r += 1
piv = short.pivot_table(
    values='req_count_satisfaction_snapp',
    index='age_label', columns='cooperation_type', aggfunc='mean'
).reindex([a for a in AGE_ORDER if a in short['age_label'].unique()])
r = write_matrix(ws3, r, 2, piv,
    'Snapp Request Count Satisfaction (1-5)', row_label='Age Group',
    bg_header=SNAPP, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)
r += 1

# ⑤ All satisfaction dimensions by activity level
section_title(ws3, r, 2,
    '⑤ SNAPP — Satisfaction Dimensions by Activity Level', 12, bg=SNAPP)
ws3.row_dimensions[r].height = 24
r += 1
act_order = ['Occasional (<1h/day)', 'Part-Time (5-20h/wk)',
             'Full-Time (8-12h/day)', 'Full-Time (>40h/wk)', 'Full-Time (>12h/day)']
piv = (short
       .groupby('activity_group')[
           ['overall_satisfaction_snapp', 'fare_satisfaction_snapp',
            'income_satisfaction_snapp', 'req_count_satisfaction_snapp']
       ].mean()
       .rename(columns={
           'overall_satisfaction_snapp': 'Overall',
           'fare_satisfaction_snapp': 'Fare',
           'income_satisfaction_snapp': 'Income',
           'req_count_satisfaction_snapp': 'Request Count',
       })
       .reindex([a for a in act_order if a in short['activity_group'].unique()])
      )
r = write_matrix(ws3, r, 2, piv,
    'Snapp Satisfaction by Activity Level (1-5)', row_label='Activity Level',
    bg_header=SNAPP, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)

print("  ✓ Sheet: Matrix_Satisfaction")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 4 — Matrix_Incentive
# ─────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet('Matrix_Incentive')
ws4.sheet_view.showGridLines = False
ws4.column_dimensions['A'].width = 3

r = 2
ws4.merge_cells(f'B{r}:N{r}')
h1(ws4.cell(r, 2), '🎁 INCENTIVE ANALYSIS — By Age Group & Driver Type', bg=NAVY, size=14)
ws4.row_dimensions[r].height = 36
r += 2

INC_LO, INC_MID, INC_HI = 'FFF8CBAD', 'FFFFD966', 'FF70AD47'  # orange→yellow→green

# ① Incentive category % by age (Snapp)
section_title(ws4, r, 2,
    '① SNAPP — Incentive Category Distribution by Age Group (%)', 12, bg=SNAPP)
ws4.row_dimensions[r].height = 24
r += 1
sub = short[short['incentive_category_snapp'].notna()]
piv = sub.pivot_table(
    values='recordID', index='age_label',
    columns='incentive_category_snapp', aggfunc='count', fill_value=0
).reindex([a for a in AGE_ORDER if a in sub['age_label'].unique()])
piv_pct = piv.div(piv.sum(axis=1), axis=0) * 100
piv_pct = piv_pct.reindex(columns=[c for c in INC_ORDER if c in piv_pct.columns])
r = write_matrix(ws4, r, 2, piv_pct,
    'Snapp Incentive Type by Age Group (%)', row_label='Age Group',
    bg_header=SNAPP, hm_lo=INC_LO, hm_mid=INC_MID, hm_hi=INC_HI)
r += 1

# ② Incentive category % by driver type (Snapp)
section_title(ws4, r, 2,
    '② SNAPP — Incentive Category by Driver Type (%)', 12, bg=SNAPP)
ws4.row_dimensions[r].height = 24
r += 1
piv = sub.pivot_table(
    values='recordID', index='cooperation_type',
    columns='incentive_category_snapp', aggfunc='count', fill_value=0
)
piv_pct = piv.div(piv.sum(axis=1), axis=0) * 100
piv_pct = piv_pct.reindex(columns=[c for c in INC_ORDER if c in piv_pct.columns])
r = write_matrix(ws4, r, 2, piv_pct,
    'Snapp Incentive Type by Driver Type (%)', row_label='Driver Type',
    bg_header=SNAPP, hm_lo=INC_LO, hm_mid=INC_MID, hm_hi=INC_HI)
r += 1

# ③ Incentive satisfaction by age × category (Snapp)
section_title(ws4, r, 2,
    '③ SNAPP — Incentive Satisfaction (1-5) by Category & Age', 12, bg=SNAPP)
ws4.row_dimensions[r].height = 24
r += 1
piv = sub.pivot_table(
    values='overall_incentive_satisfaction_snapp',
    index='age_label', columns='incentive_category_snapp', aggfunc='mean'
).reindex([a for a in AGE_ORDER if a in sub['age_label'].unique()])
piv = piv.reindex(columns=[c for c in INC_ORDER if c in piv.columns])
r = write_matrix(ws4, r, 2, piv,
    'Snapp Incentive Satisfaction by Age & Category (1-5)', row_label='Age Group',
    bg_header=SNAPP, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)
r += 1

# ④ Incentive satisfaction by driver type × category (Snapp)
section_title(ws4, r, 2,
    '④ SNAPP — Incentive Satisfaction (1-5) by Category & Driver Type', 12, bg=SNAPP)
ws4.row_dimensions[r].height = 24
r += 1
piv = sub.pivot_table(
    values='overall_incentive_satisfaction_snapp',
    index='cooperation_type', columns='incentive_category_snapp', aggfunc='mean'
)
piv = piv.reindex(columns=[c for c in INC_ORDER if c in piv.columns])
r = write_matrix(ws4, r, 2, piv,
    'Snapp Incentive Satisfaction by Driver Type & Category (1-5)', row_label='Driver Type',
    bg_header=SNAPP, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)
r += 1

# ⑤ Detailed incentive types (Pay After Ride, Commission-Free, etc.) by age
section_title(ws4, r, 2,
    '⑤ SNAPP — Incentive Type Details by Age Group (%)', 12, bg=SNAPP)
ws4.row_dimensions[r].height = 24
r += 1
inc_types = (
    long[long['main_question'] == 'incentive_type_snapp'][['recordID', 'answer']]
    .merge(short[['recordID', 'age_label']], on='recordID', how='left')
)
piv = inc_types.pivot_table(
    values='recordID', index='age_label', columns='answer', aggfunc='count', fill_value=0
).reindex([a for a in AGE_ORDER if a in inc_types['age_label'].unique()])
piv_pct = piv.div(piv.sum(axis=1), axis=0) * 100
r = write_matrix(ws4, r, 2, piv_pct,
    'Snapp Incentive Type Detail by Age Group (%)', row_label='Age Group',
    bg_header=SNAPP, hm_lo=INC_LO, hm_mid=INC_MID, hm_hi=INC_HI)

print("  ✓ Sheet: Matrix_Incentive")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 5 — Matrix_Behavior
# ─────────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet('Matrix_Behavior')
ws5.sheet_view.showGridLines = False
ws5.column_dimensions['A'].width = 3

r = 2
ws5.merge_cells(f'B{r}:N{r}')
h1(ws5.cell(r, 2),
   '🚗 DRIVER BEHAVIOR ANALYSIS — Demand, NPS, Decline Reasons',
   bg=NAVY, size=14)
ws5.row_dimensions[r].height = 36
r += 2

short['recommend_snapp_n'] = pd.to_numeric(short['recommend_snapp'], errors='coerce')

# ① NPS Snapp by age × driver type
section_title(ws5, r, 2,
    '① SNAPP — NPS Score (0-10) by Age Group & Driver Type', 12, bg=SNAPP)
ws5.row_dimensions[r].height = 24
r += 1
piv = short.pivot_table(
    values='recommend_snapp_n', index='age_label',
    columns='cooperation_type', aggfunc='mean'
).reindex([a for a in AGE_ORDER if a in short['age_label'].unique()])
r = write_matrix(ws5, r, 2, piv,
    'Snapp NPS by Age & Driver Type (Avg 0-10)', row_label='Age Group',
    bg_header=SNAPP, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)
r += 1

# ② Demand acceptance rate distribution by age
section_title(ws5, r, 2,
    '② SNAPP — Demand Acceptance Rate by Age Group (%)', 12, bg=SNAPP)
ws5.row_dimensions[r].height = 24
r += 1
dem_order = ['less than half', 'almost half', 'more than half', 'almost all']
piv = short.pivot_table(
    values='recordID', index='age_label',
    columns='demand_process', aggfunc='count', fill_value=0
).reindex([a for a in AGE_ORDER if a in short['age_label'].unique()])
piv_pct = piv.div(piv.sum(axis=1), axis=0) * 100
piv_pct = piv_pct.reindex(columns=[c for c in dem_order if c in piv_pct.columns])
r = write_matrix(ws5, r, 2, piv_pct,
    'Demand Acceptance Rate Distribution by Age (%)', row_label='Age Group',
    bg_header=SNAPP, hm_lo=INC_LO, hm_mid=INC_MID, hm_hi=INC_HI)
r += 1

# ③ Decline reasons by age
section_title(ws5, r, 2,
    '③ SNAPP — Decline Reasons by Age Group (% selecting each reason)', 12, bg=SNAPP)
ws5.row_dimensions[r].height = 24
r += 1
decline_df = (
    long[long['main_question'] == 'decline_reason'][['recordID', 'sub_question']]
    .merge(short[['recordID', 'age_label']], on='recordID', how='left')
)
decline_df['reason'] = (decline_df['sub_question']
    .str.replace('decline_reason_', '', regex=False)
    .str.replace('_', ' ')
    .str.title())
total_per_age = short.groupby('age_label')['recordID'].count()
piv = decline_df.pivot_table(
    values='recordID', index='age_label', columns='reason',
    aggfunc='count', fill_value=0
).reindex([a for a in AGE_ORDER if a in decline_df['age_label'].unique()])
piv_pct = piv.div(total_per_age.reindex(piv.index), axis=0) * 100
r = write_matrix(ws5, r, 2, piv_pct,
    'Decline Reasons by Age Group (% of drivers in that age)', row_label='Age Group',
    bg_header=SNAPP, hm_lo=INC_LO, hm_mid=INC_MID, hm_hi=INC_HI)
r += 1

# ④ Decline reasons by driver type
section_title(ws5, r, 2,
    '④ SNAPP — Decline Reasons by Driver Type (% selecting each reason)', 12, bg=SNAPP)
ws5.row_dimensions[r].height = 24
r += 1
decline_df2 = (
    long[long['main_question'] == 'decline_reason'][['recordID', 'sub_question']]
    .merge(short[['recordID', 'cooperation_type']], on='recordID', how='left')
)
decline_df2['reason'] = (decline_df2['sub_question']
    .str.replace('decline_reason_', '', regex=False)
    .str.replace('_', ' ')
    .str.title())
total_per_type = short.groupby('cooperation_type')['recordID'].count()
piv = decline_df2.pivot_table(
    values='recordID', index='cooperation_type', columns='reason',
    aggfunc='count', fill_value=0
)
piv_pct = piv.div(total_per_type.reindex(piv.index), axis=0) * 100
r = write_matrix(ws5, r, 2, piv_pct,
    'Decline Reasons by Driver Type (% of drivers)', row_label='Driver Type',
    bg_header=SNAPP, hm_lo=INC_LO, hm_mid=INC_MID, hm_hi=INC_HI)

print("  ✓ Sheet: Matrix_Behavior")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 6 — Matrix_Tapsi  (dual-platform drivers only)
# ─────────────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet('Matrix_Tapsi')
ws6.sheet_view.showGridLines = False
ws6.column_dimensions['A'].width = 3

r = 2
ws6.merge_cells(f'B{r}:N{r}')
h1(ws6.cell(r, 2),
   f'🔴 TAPSI ANALYSIS — Drivers Using Both Platforms (n={len(tapsi):,})',
   bg=TAPSI, size=14)
ws6.row_dimensions[r].height = 36
r += 2

# ① Tapsi vs Snapp satisfaction comparison by driver type
section_title(ws6, r, 2,
    '① TAPSI vs SNAPP — Satisfaction Comparison by Driver Type (Avg 1-5)', 12, bg=TAPSI)
ws6.row_dimensions[r].height = 24
r += 1
comp = tapsi.groupby('cooperation_type').agg(
    snapp_overall=('overall_satisfaction_snapp', 'mean'),
    tapsi_overall=('overall_satisfaction_tapsi', 'mean'),
    snapp_fare=('fare_satisfaction_snapp', 'mean'),
    tapsi_fare=('fare_satisfaction_tapsi', 'mean'),
    snapp_income=('income_satisfaction_snapp', 'mean'),
    tapsi_income=('income_satisfaction_tapsi', 'mean'),
    snapp_req=('req_count_satisfaction_snapp', 'mean'),
    tapsi_req=('req_count_satisfaction_tapsi', 'mean'),
).T
comp.index = ['Snapp Overall', 'Tapsi Overall', 'Snapp Fare', 'Tapsi Fare',
              'Snapp Income', 'Tapsi Income', 'Snapp Req.Count', 'Tapsi Req.Count']
r = write_matrix(ws6, r, 2, comp,
    'Tapsi vs Snapp Satisfaction by Driver Type (1-5)', row_label='Metric',
    bg_header=TAPSI, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)
r += 1

# ② Tapsi overall satisfaction by age × driver type
section_title(ws6, r, 2,
    '② TAPSI — Overall Satisfaction by Age Group & Driver Type', 12, bg=TAPSI)
ws6.row_dimensions[r].height = 24
r += 1
piv = tapsi.pivot_table(
    values='overall_satisfaction_tapsi', index='age_label',
    columns='cooperation_type', aggfunc='mean'
).reindex([a for a in AGE_ORDER if a in tapsi['age_label'].unique()])
r = write_matrix(ws6, r, 2, piv,
    'Tapsi Overall Satisfaction by Age & Driver Type (1-5)', row_label='Age Group',
    bg_header=TAPSI, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)
r += 1

# ③ Tapsi incentive category % by age
section_title(ws6, r, 2,
    '③ TAPSI — Incentive Category Distribution by Age Group (%)', 12, bg=TAPSI)
ws6.row_dimensions[r].height = 24
r += 1
sub_t = tapsi[tapsi['incentive_category_tapsi'].notna()]
piv = sub_t.pivot_table(
    values='recordID', index='age_label',
    columns='incentive_category_tapsi', aggfunc='count', fill_value=0
).reindex([a for a in AGE_ORDER if a in sub_t['age_label'].unique()])
piv_pct = piv.div(piv.sum(axis=1), axis=0) * 100
piv_pct = piv_pct.reindex(columns=[c for c in INC_ORDER if c in piv_pct.columns])
r = write_matrix(ws6, r, 2, piv_pct,
    'Tapsi Incentive Type by Age Group (%)', row_label='Age Group',
    bg_header=TAPSI, hm_lo=INC_LO, hm_mid=INC_MID, hm_hi=INC_HI)
r += 1

# ④ Tapsi incentive satisfaction by age × category
section_title(ws6, r, 2,
    '④ TAPSI — Incentive Satisfaction (1-5) by Age & Category', 12, bg=TAPSI)
ws6.row_dimensions[r].height = 24
r += 1
piv = sub_t.pivot_table(
    values='overall_incentive_satisfaction_tapsi',
    index='age_label', columns='incentive_category_tapsi', aggfunc='mean'
).reindex([a for a in AGE_ORDER if a in sub_t['age_label'].unique()])
piv = piv.reindex(columns=[c for c in INC_ORDER if c in piv.columns])
r = write_matrix(ws6, r, 2, piv,
    'Tapsi Incentive Satisfaction by Age & Category (1-5)', row_label='Age Group',
    bg_header=TAPSI, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)
r += 1

# ⑤ Tapsi NPS by age × driver type
section_title(ws6, r, 2,
    '⑤ TAPSI — NPS Score (0-10) by Age Group & Driver Type', 12, bg=TAPSI)
ws6.row_dimensions[r].height = 24
r += 1
tapsi['recommend_tapsi_n'] = pd.to_numeric(tapsi['recommend_tapsi'], errors='coerce')
piv = tapsi.pivot_table(
    values='recommend_tapsi_n', index='age_label',
    columns='cooperation_type', aggfunc='mean'
).reindex([a for a in AGE_ORDER if a in tapsi['age_label'].unique()])
r = write_matrix(ws6, r, 2, piv,
    'Tapsi NPS by Age & Driver Type (Avg 0-10)', row_label='Age Group',
    bg_header=TAPSI, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)

print("  ✓ Sheet: Matrix_Tapsi")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 7 — Summary_Stats  (KPI dashboard)
# ─────────────────────────────────────────────────────────────────────────────
ws7 = wb.create_sheet('Summary_Stats')
ws7.sheet_view.showGridLines = False
ws7.column_dimensions['A'].width = 3
ws7.column_dimensions['B'].width = 38
ws7.column_dimensions['C'].width = 24
ws7.column_dimensions['D'].width = 24
ws7.column_dimensions['E'].width = 24

r = 2
ws7.merge_cells(f'B{r}:E{r}')
h1(ws7.cell(r, 2), '📈 SURVEY SUMMARY — Key Performance Indicators', bg=NAVY, size=14)
ws7.row_dimensions[r].height = 36
r += 2

n = len(short)
n_pt = (short['cooperation_type'] == 'Part-Time').sum()
n_ft = (short['cooperation_type'] == 'Full-Time').sum()
n_snapp_only = (short['has_tapsi'] == 'Snapp Only').sum()
n_both = len(tapsi)

def avg(col, df=short):
    return df[col].mean()

def avg_str(col, df=short):
    return f"{avg(col, df):.2f}"

def avg_by(col, by_col, by_val, df=short):
    return f"{df[df[by_col]==by_val][col].mean():.2f}"

kpis = [
    # (label, value_col1, value_col2, value_col3)
    # section headers have empty value cols
    ('SAMPLE SIZE', '', '', ''),
    ('Total Respondents',   f'{n:,}',
     f'Snapp-only: {n_snapp_only:,}', f'Used Both: {n_both:,}'),
    ('Part-Time Drivers',   f'{n_pt:,}',
     f'{n_pt/n*100:.1f}% of total', ''),
    ('Full-Time Drivers',   f'{n_ft:,}',
     f'{n_ft/n*100:.1f}% of total', ''),
    ('', '', '', ''),
    ('SNAPP SATISFACTION (ALL DRIVERS)', '', '', ''),
    ('Overall Satisfaction (1-5)',  avg_str('overall_satisfaction_snapp'),
     f'Part-Time: {avg_by("overall_satisfaction_snapp","cooperation_type","Part-Time")}',
     f'Full-Time: {avg_by("overall_satisfaction_snapp","cooperation_type","Full-Time")}'),
    ('Fare Satisfaction (1-5)',     avg_str('fare_satisfaction_snapp'),
     f'Part-Time: {avg_by("fare_satisfaction_snapp","cooperation_type","Part-Time")}',
     f'Full-Time: {avg_by("fare_satisfaction_snapp","cooperation_type","Full-Time")}'),
    ('Income Satisfaction (1-5)',   avg_str('income_satisfaction_snapp'),
     f'Part-Time: {avg_by("income_satisfaction_snapp","cooperation_type","Part-Time")}',
     f'Full-Time: {avg_by("income_satisfaction_snapp","cooperation_type","Full-Time")}'),
    ('Request Satisfaction (1-5)',  avg_str('req_count_satisfaction_snapp'),
     f'Part-Time: {avg_by("req_count_satisfaction_snapp","cooperation_type","Part-Time")}',
     f'Full-Time: {avg_by("req_count_satisfaction_snapp","cooperation_type","Full-Time")}'),
    ('Incentive Satisfaction (1-5)',avg_str('overall_incentive_satisfaction_snapp'), '', ''),
    ('Avg NPS (0-10)',
     f'{pd.to_numeric(short["recommend_snapp"], errors="coerce").mean():.2f}', '', ''),
    ('', '', '', ''),
    ('TAPSI SATISFACTION (DUAL-PLATFORM DRIVERS)', '', '', ''),
    ('Overall Satisfaction (1-5)',  avg_str('overall_satisfaction_tapsi', tapsi),
     f'Part-Time: {avg_by("overall_satisfaction_tapsi","cooperation_type","Part-Time",tapsi)}',
     f'Full-Time: {avg_by("overall_satisfaction_tapsi","cooperation_type","Full-Time",tapsi)}'),
    ('Fare Satisfaction (1-5)',     avg_str('fare_satisfaction_tapsi', tapsi), '', ''),
    ('Income Satisfaction (1-5)',   avg_str('income_satisfaction_tapsi', tapsi), '', ''),
    ('Avg NPS (0-10)',
     f'{pd.to_numeric(tapsi["recommend_tapsi"], errors="coerce").mean():.2f}', '', ''),
    ('', '', '', ''),
    ('INCENTIVES — SNAPP', '', '', ''),
    ('Received Incentive Message',
     f'{(short["incentive_got_message_snapp"]=="Yes").sum():,}',
     f'{(short["incentive_got_message_snapp"]=="Yes").sum()/n*100:.1f}% of drivers', ''),
    ('Money Incentive',
     f'{(short["incentive_category_snapp"]=="Money").sum():,}', '', ''),
    ('Free-Commission',
     f'{(short["incentive_category_snapp"]=="Free-Commission").sum():,}', '', ''),
    ('Money & Free-Commission',
     f'{(short["incentive_category_snapp"]=="Money & Free-commission").sum():,}', '', ''),
    ('', '', '', ''),
    ('DEMAND & BEHAVIOR', '', '', ''),
    ('Accept "Almost All" Requests',
     f'{(short["demand_process"]=="almost all").sum():,}',
     f'{(short["demand_process"]=="almost all").sum()/n*100:.1f}%', ''),
    ('Top Decline Reason',
     'Fare (most common)', '', ''),
    ('Avg Missed Demands per 10',
     f'{pd.to_numeric(short["missed_demand_per_10"], errors="coerce").mean():.2f}', '', ''),
]

for i, (label, v1, v2, v3) in enumerate(kpis):
    rr = r + i
    ws7.row_dimensions[rr].height = 22
    is_section_header = label and not v1
    if is_section_header:
        ws7.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=5)
        h2(ws7.cell(rr, 2), label, bg=BLUE)
    else:
        bg = LGRAY if i % 2 == 0 else WHITE
        ws7.cell(rr, 2).value = label
        ws7.cell(rr, 2).font = Font(size=10, name='Calibri')
        ws7.cell(rr, 2).fill = PatternFill('solid', start_color=bg)
        ws7.cell(rr, 2).alignment = Alignment(vertical='center')
        for j, v in enumerate([v1, v2, v3]):
            cell = ws7.cell(rr, 3 + j)
            cell.value = v
            cell.font = Font(size=10, name='Calibri', bold=(j == 0))
            cell.fill = PatternFill('solid', start_color=bg)
            cell.alignment = Alignment(horizontal='center', vertical='center')

print("  ✓ Sheet: Summary_Stats")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 8 — Matrix_CrossPlatform  (Snapp vs Tapsi, dual-platform drivers only)
# ─────────────────────────────────────────────────────────────────────────────
ws8 = wb.create_sheet('Matrix_CrossPlatform')
ws8.sheet_view.showGridLines = False
ws8.column_dimensions['A'].width = 3

r = 2
ws8.merge_cells(f'B{r}:Q{r}')
h1(ws8.cell(r, 2),
   '⚖️  SNAPP vs TAPSI — Side-by-Side (Dual-Platform Drivers Only)',
   bg=NAVY, size=14)
ws8.row_dimensions[r].height = 36
r += 2

# ① Snapp overall satisfaction (dual-platform)
section_title(ws8, r, 2,
    '① SNAPP Overall Satisfaction — by Age × Driver Type', 14, bg=SNAPP)
ws8.row_dimensions[r].height = 24
r += 1
piv = tapsi.pivot_table(
    values='overall_satisfaction_snapp', index='age_label',
    columns='cooperation_type', aggfunc='mean'
).reindex([a for a in AGE_ORDER if a in tapsi['age_label'].unique()])
r = write_matrix(ws8, r, 2, piv,
    'Snapp Overall Satisfaction — dual-platform drivers (1-5)', row_label='Age Group',
    bg_header=SNAPP, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)

# ② Tapsi overall satisfaction (dual-platform)
section_title(ws8, r, 2,
    '② TAPSI Overall Satisfaction — by Age × Driver Type', 14, bg=TAPSI)
ws8.row_dimensions[r].height = 24
r += 1
piv = tapsi.pivot_table(
    values='overall_satisfaction_tapsi', index='age_label',
    columns='cooperation_type', aggfunc='mean'
).reindex([a for a in AGE_ORDER if a in tapsi['age_label'].unique()])
r = write_matrix(ws8, r, 2, piv,
    'Tapsi Overall Satisfaction — dual-platform drivers (1-5)', row_label='Age Group',
    bg_header=TAPSI, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)

# ③ Snapp incentive satisfaction (dual-platform)
section_title(ws8, r, 2,
    '③ SNAPP Incentive Satisfaction — by Age × Incentive Type', 14, bg=SNAPP)
ws8.row_dimensions[r].height = 24
r += 1
sub = tapsi[tapsi['incentive_category_snapp'].notna()]
piv = sub.pivot_table(
    values='overall_incentive_satisfaction_snapp',
    index='age_label', columns='incentive_category_snapp', aggfunc='mean'
).reindex([a for a in AGE_ORDER if a in sub['age_label'].unique()])
piv = piv.reindex(columns=[c for c in INC_ORDER if c in piv.columns])
r = write_matrix(ws8, r, 2, piv,
    'Snapp Incentive Satisfaction — dual-platform (1-5)', row_label='Age Group',
    bg_header=SNAPP, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)

# ④ Tapsi incentive satisfaction (dual-platform)
section_title(ws8, r, 2,
    '④ TAPSI Incentive Satisfaction — by Age × Incentive Type', 14, bg=TAPSI)
ws8.row_dimensions[r].height = 24
r += 1
sub_t = tapsi[tapsi['incentive_category_tapsi'].notna()]
piv = sub_t.pivot_table(
    values='overall_incentive_satisfaction_tapsi',
    index='age_label', columns='incentive_category_tapsi', aggfunc='mean'
).reindex([a for a in AGE_ORDER if a in sub_t['age_label'].unique()])
piv = piv.reindex(columns=[c for c in INC_ORDER if c in piv.columns])
r = write_matrix(ws8, r, 2, piv,
    'Tapsi Incentive Satisfaction — dual-platform (1-5)', row_label='Age Group',
    bg_header=TAPSI, hm_lo=SAT_LO, hm_mid=SAT_MID, hm_hi=SAT_HI)

print("  ✓ Sheet: Matrix_CrossPlatform")

# ══════════════════════════════════════════════════════════════════════════════
# TAB COLORS & SAVE
# ══════════════════════════════════════════════════════════════════════════════

ws0.sheet_properties.tabColor = '1F3864'
ws1.sheet_properties.tabColor = '2E75B6'
ws2.sheet_properties.tabColor = '833D11'
ws3.sheet_properties.tabColor = '1A5276'
ws4.sheet_properties.tabColor = '1E8449'
ws5.sheet_properties.tabColor = '6C3483'
ws6.sheet_properties.tabColor = '922B21'
ws7.sheet_properties.tabColor = 'F39C12'
ws8.sheet_properties.tabColor = '117A65'

wb.save(OUTPUT_XLSX)
print(f"\n✅ Done! Output saved to: {OUTPUT_XLSX}")
print(f"   Open in Excel or import directly into Power BI Desktop.")
